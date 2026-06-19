#!/usr/bin/env python3
"""Grade model-test output workbooks with one or more judge models."""

from __future__ import annotations

import argparse
import concurrent.futures
import csv
import datetime as dt
import json
import re
import sys
import threading
import time
import traceback
import zipfile
from collections import Counter
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment
except ImportError as exc:  # pragma: no cover - exercised by user environment
    print(
        "Missing dependency: openpyxl. Install it with "
        "`python3 -m pip install openpyxl`.",
        file=sys.stderr,
    )
    raise SystemExit(2) from exc

from model_eval_runner import (
    RateLimitError,
    average,
    call_openai_compatible,
    clean_text,
    default_config,
    enabled_models,
    find_header_row,
    group_pairs_by_product,
    model_capabilities,
    normalized_header,
    parse_weight,
    product_lane_label,
    product_worker_count,
    read_model_workbook,
    read_scored_model_keys,
    save_workbook_atomic,
    set_widths,
    split_filter,
    stringify_usage,
    style_header,
    usage_token_count,
    usage_detail_token_count,
    validate_api_keys,
    validate_openrouter_model_ids,
)


DEFAULT_MAX_TOKENS = 800
RESULT_SHEET_NAMES = {"Run Summary", "Run Results", "Skipped Tests"}
GRADE_SHEET_NAMES = {
    "Grade Summary",
    "Model Scores",
    "AI Scores",
    "Consensus",
    "Results",
    "Grades",
    "Skipped Outputs",
}
PREFERRED_RUBRIC_SHEET_NAMES = (
    "Scoring Guide",
    "Rubric",
    "Rubric v3",
    "Model Testing Rubric",
)
PREFERRED_WEIGHT_SHEET_NAMES = ("Weights", "Category Weights")
SCORE_CLEARING_SKIPPED_REASONS = {"missing source output", "source output errored"}

RESULT_HEADER_ALIASES = {
    "source_run_id": ["run_id", "Run ID", "Source Run ID"],
    "timestamp": ["timestamp", "Timestamp"],
    "source_model_key": ["model_key", "Model Key", "Source Model Key", "Tested Model Key"],
    "source_model_name": ["model_name", "Model Name", "Source Model Name", "Tested Model Name"],
    "source_provider": ["provider", "Provider", "Source Provider"],
    "source_provider_model": ["provider_model", "Provider Model", "Source Provider Model"],
    "test_id": ["test_id", "TESTID", "Test ID", "TestID", "Prompt ID", "ID"],
    "category": ["category", "Category"],
    "criterion": ["criterion", "Criterion"],
    "weight": ["weight", "Weight"],
    "eval_method": ["eval_method", "Eval Method", "Output Type", "Test Type"],
    "output_type": ["output_type", "Output Type", "Response Type"],
    "response": ["response", "Response", "Model Response", "Output", "Answer"],
    "output_files": ["output_files", "Output Files", "Files"],
    "output_urls": ["output_urls", "Output URLs", "URLs"],
    "source_score": ["score", "Score", "Original Score"],
    "source_reasoning": ["reasoning", "Reasoning", "Original Reasoning"],
    "error": ["error", "Error"],
}

PROMPT_HEADER_ALIASES = {
    "test_id": ["TESTID", "Test ID", "TestID", "Prompt ID", "ID"],
    "category": ["Category"],
    "criterion": ["Criterion"],
    "weight": ["Weight"],
    "eval_method": ["Eval Method", "Type", "Test Type", "Prompt Type", "Output Type"],
    "prompt": ["Prompt", "Benchmark Prompt", "Original Prompt"],
    "input_material": [
        "Additional source information",
        "Additional Source Information",
        "Additional Source Info",
        "Source Information",
        "Input Material",
    ],
}

RUBRIC_HEADER_ALIASES = {
    "test_id": ["TESTID", "Test ID", "TestID", "Prompt ID", "ID"],
    "category": ["Category"],
    "criterion": ["Criterion"],
    "website_field": ["Website Field", "Rails Field", "Ruby Field", "Score Field", "DB Field"],
    "weight": ["Weight"],
    "prompt": ["Prompt", "Benchmark Prompt", "Original Prompt"],
    "input_material": [
        "Additional source information",
        "Additional Source Information",
        "Additional Source Info",
        "Source Information",
        "Input Material",
    ],
    "rubric": [
        "Rubric",
        "Scoring Rubric",
        "Scoring Guidance",
        "Score Guidance",
        "Grading Instructions",
        "Judge Instructions",
        "Evaluation Rubric",
        "Criteria",
    ],
    "what_it_measures": [
        "What it measures",
        "What it tests",
        "What this measures",
        "Measure",
        "Description",
    ],
    "score_1_3": ["1-3 (Poor)", "1–3 (Poor)", "1 to 3 (Poor)", "Poor", "1-3"],
    "score_4_6": [
        "4-6 (Adequate)",
        "4–6 (Adequate)",
        "4 to 6 (Adequate)",
        "Adequate",
        "4-6",
    ],
    "score_7_8": ["7-8 (Strong)", "7–8 (Strong)", "7 to 8 (Strong)", "Strong", "7-8"],
    "score_9_10": [
        "9-10 (Excellent)",
        "9–10 (Excellent)",
        "9 to 10 (Excellent)",
        "Excellent",
        "9-10",
    ],
    "score_min": ["Minimum Score", "Min Score", "Score Min"],
    "score_max": ["Maximum Score", "Max Score", "Score Max"],
    "enabled": ["Enabled", "Run", "Include"],
}

CATEGORY_WEIGHT_HEADER_ALIASES = {
    "category": ["Category", "Score Category"],
    "weight": ["Category Weight", "Weight", "Overall Weight"],
}

GRADE_COLUMNS = [
    "grading_run_id",
    "timestamp",
    "source_row",
    "source_run_id",
    "source_model_key",
    "source_model_name",
    "source_provider",
    "source_provider_model",
    "test_id",
    "category",
    "criterion",
    "weight",
    "grader_model_key",
    "grader_model_name",
    "grader_provider",
    "grader_provider_model",
    "score",
    "reasoning",
    "strengths",
    "issues",
    "rubric",
    "prompt",
    "input_material",
    "source_response",
    "source_output_files",
    "source_output_urls",
    "source_score",
    "source_reasoning",
    "latency_seconds",
    "prompt_tokens",
    "completion_tokens",
    "reasoning_tokens",
    "total_tokens",
    "usage",
    "error",
]

WEBSITE_DIRECT_FIELD_TEST_IDS = {
    "write_edit_score": ["W1"],
    "score_write_edit": ["W1"],
    "summarisation_score": ["W2"],
    "score_summarization": ["W2"],
    "research_fact_checking_score": ["R1"],
    "score_research_fact_check": ["R1"],
    "source_quality_score": ["R2", "A&2"],
    "score_source_quality": ["R2", "A&2"],
    "hallucination_resistance_score": ["R3", "A&1"],
    "score_hallucination_resistance": ["R3", "A&1"],
    "deep_research_score": ["R4"],
    "coding_speed_score": ["C1"],
    "score_coding_speed": ["C1"],
    "coding_accuracy_score": ["C2"],
    "debugging_score": ["C3"],
    "agentic_coding_score": ["C4"],
    "consistency_score": ["A&3"],
    "score_consistency": ["A&3"],
    "reasoning_score": ["A&4"],
    "score_logic": ["A&4"],
    "image_quality_score": ["IG1"],
    "prompt_adherence_score": ["IG2"],
    "text_rendering_score": ["IG3"],
    "image_editing_score": ["IG4"],
    "transcription_score": ["M1"],
    "score_meetings_transcription": ["M1"],
    "meeting_summary_score": ["M2"],
    "follow_up_score": ["M3"],
    "translation_accuracy_score": ["T1"],
    "score_translation_accuracy": ["T1"],
    "translation_speed_score": ["T2"],
    "score_translation_speed": ["T2"],
}

WEBSITE_COMPOSITE_FIELD_TEST_IDS = {
    "score_text_generation": ["W1", "W2", "W3"],
    "score_email_writing": ["W1"],
    "score_coding": ["C1", "C2", "C3", "C4"],
    "score_image_generation": ["IG1", "IG2", "IG3", "IG4"],
    "score_accuracy": ["A&1", "A&2", "A&3", "A&4"],
}


@dataclass(frozen=True)
class PromptInfo:
    test_id: str
    category: str = ""
    criterion: str = ""
    weight: float = 1.0
    eval_method: str = ""
    prompt: str = ""
    input_material: str = ""


@dataclass(frozen=True)
class TestOutput:
    row_number: int
    output_key: str
    source_run_id: str
    source_model_key: str
    source_model_name: str
    source_provider: str
    source_provider_model: str
    test_id: str
    category: str
    criterion: str
    weight: float
    eval_method: str
    output_type: str
    response: str
    output_files: str
    output_urls: str
    source_score: str
    source_reasoning: str
    prompt: str
    input_material: str
    error: str


@dataclass(frozen=True)
class RubricEntry:
    row_number: int
    test_id: str
    category: str
    criterion: str
    website_field: str
    weight: float
    prompt: str
    input_material: str
    rubric: str
    score_min: float
    score_max: float


@dataclass(frozen=True)
class RubricContext:
    rubric: str
    score_min: float
    score_max: float
    weight: float
    category: str
    criterion: str
    prompt: str
    input_material: str
    matched: bool


@dataclass(frozen=True)
class SkippedOutput:
    row_number: int
    test_id: str
    source_model_key: str
    reason: str


def cell_value(ws: Any, row_number: int, columns: dict[str, int], field: str) -> Any:
    column = columns.get(field)
    return ws.cell(row_number, column).value if column else ""


def truthy(value: Any, default: bool = True) -> bool:
    text = clean_text(value).lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "y", "run", "include", "enabled"}


def parse_score_bound(value: Any, default: float) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def parse_weight_with_fallback(value: Any, fallback: float) -> float:
    if value in (None, ""):
        return fallback
    return parse_weight(value)


def normalize_key(value: str) -> str:
    return clean_text(value).strip().lower()


def is_xlsx_workbook(path: Path) -> bool:
    return zipfile.is_zipfile(path)


def find_csv_columns(
    headers: list[str],
    aliases: dict[str, list[str]],
    required_fields: set[str],
) -> dict[str, str]:
    normalized_values = {
        normalized_header(header): header
        for header in headers
        if clean_text(header)
    }

    positions: dict[str, str] = {}
    for field, field_aliases in aliases.items():
        for alias in field_aliases:
            header = normalized_values.get(normalized_header(alias))
            if header:
                positions[field] = header
                break

    missing = sorted(required_fields - positions.keys())
    if missing:
        raise ValueError(
            f"Could not find CSV header columns: {', '.join(missing)}. "
            f"Found: {', '.join(headers)}"
        )
    return positions


def csv_value(row: dict[str, str], columns: dict[str, str], field: str) -> Any:
    header = columns.get(field)
    return row.get(header, "") if header else ""


def canonical_identity(value: Any) -> set[str]:
    text = clean_text(value).lower()
    if not text:
        return set()

    identities = {re.sub(r"[^a-z0-9]+", "", text)}
    if "/" in text:
        identities.add(re.sub(r"[^a-z0-9]+", "", text.rsplit("/", 1)[-1]))
    return {identity for identity in identities if identity}


def grader_identities(grader: dict[str, Any]) -> set[str]:
    identities: set[str] = set()
    for field in ("key", "name", "model"):
        identities.update(canonical_identity(grader.get(field)))
    return identities


def output_identities(output: TestOutput) -> set[str]:
    identities: set[str] = set()
    for value in (
        output.source_model_key,
        output.source_model_name,
        output.source_provider_model,
    ):
        identities.update(canonical_identity(value))
    return identities


def row_grader_identities(row: dict[str, Any]) -> set[str]:
    identities: set[str] = set()
    for field in ("grader_model_key", "grader_model_name", "grader_provider_model"):
        identities.update(canonical_identity(row.get(field)))
    return identities


def row_source_identities(row: dict[str, Any]) -> set[str]:
    identities: set[str] = set()
    for field in ("source_model_key", "source_model_name", "source_provider_model"):
        identities.update(canonical_identity(row.get(field)))
    return identities


def is_self_judge(grader: dict[str, Any], output: TestOutput) -> bool:
    return bool(grader_identities(grader) & output_identities(output))


def is_self_judge_row(row: dict[str, Any]) -> bool:
    return bool(row_grader_identities(row) & row_source_identities(row))


def category_weight_for(category: str, category_weights: dict[str, float]) -> float:
    if not category:
        return 1.0
    if category in category_weights:
        return category_weights[category]
    normalized = normalize_key(category)
    for weighted_category, weight in category_weights.items():
        if normalize_key(weighted_category) == normalized:
            return weight
    return 1.0


def round_half_up(value: float) -> int:
    rounded = int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    return min(10, max(1, rounded))


def extract_json_object(text: str) -> dict[str, Any]:
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    fenced = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, flags=re.DOTALL)
    if fenced:
        return json.loads(fenced.group(1))

    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        return json.loads(text[start : end + 1])

    raise ValueError("No JSON object found in grader response.")


def selected_sheet(wb: Any, requested_name: str | None, preferred_name: str) -> Any:
    if requested_name:
        if requested_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet {requested_name!r} not found. Available sheets: {', '.join(wb.sheetnames)}"
            )
        return wb[requested_name]
    if preferred_name in wb.sheetnames:
        return wb[preferred_name]
    return wb[wb.sheetnames[0]]


def find_rubric_header_row(ws: Any) -> tuple[int, dict[str, int]]:
    try:
        return find_header_row(ws, RUBRIC_HEADER_ALIASES, {"rubric"})
    except ValueError:
        pass

    try:
        return find_header_row(ws, RUBRIC_HEADER_ALIASES, {"what_it_measures"})
    except ValueError as exc:
        raise ValueError(
            "The rubric workbook needs either a Rubric/Scoring Guidance column "
            "or a score-band layout with a What it measures column."
        ) from exc


def rubric_sheet_and_header(
    wb: Any,
    requested_name: str | None,
) -> tuple[Any, int, dict[str, int]]:
    if requested_name:
        ws = selected_sheet(wb, requested_name, wb.sheetnames[0])
        header_row, columns = find_rubric_header_row(ws)
        return ws, header_row, columns

    candidates: list[Any] = []
    for sheet_name in PREFERRED_RUBRIC_SHEET_NAMES:
        if sheet_name in wb.sheetnames:
            candidates.append(wb[sheet_name])
    candidates.extend(ws for ws in wb.worksheets if ws not in candidates)

    for ws in candidates:
        if ws.title in RESULT_SHEET_NAMES or ws.title in GRADE_SHEET_NAMES:
            continue
        try:
            header_row, columns = find_rubric_header_row(ws)
            return ws, header_row, columns
        except ValueError:
            continue

    raise ValueError(
        "The rubric workbook needs a sheet with a header row containing "
        "Test ID plus either Rubric/Scoring Guidance or What it measures."
    )


def read_category_weights(workbook_path: Path) -> dict[str, float]:
    wb = load_workbook(workbook_path, data_only=True)
    candidates: list[Any] = []
    for sheet_name in PREFERRED_WEIGHT_SHEET_NAMES:
        if sheet_name in wb.sheetnames:
            candidates.append(wb[sheet_name])
    candidates.extend(ws for ws in wb.worksheets if ws not in candidates)

    for ws in candidates:
        try:
            header_row, columns = find_header_row(
                ws,
                CATEGORY_WEIGHT_HEADER_ALIASES,
                {"category", "weight"},
            )
        except ValueError:
            continue

        weights: dict[str, float] = {}
        for row_number in range(header_row + 1, ws.max_row + 1):
            category = clean_text(cell_value(ws, row_number, columns, "category"))
            if not category:
                continue
            if normalize_key(category) in {"total", "overall", "grand total"}:
                continue
            weights[category] = parse_weight(cell_value(ws, row_number, columns, "weight"))
        if weights:
            return weights

    return {}


def composed_rubric_text(ws: Any, row_number: int, columns: dict[str, int]) -> str:
    direct = clean_text(cell_value(ws, row_number, columns, "rubric"))
    if direct:
        return direct

    measures = clean_text(cell_value(ws, row_number, columns, "what_it_measures"))
    bands = [
        ("1-3 Poor", clean_text(cell_value(ws, row_number, columns, "score_1_3"))),
        ("4-6 Adequate", clean_text(cell_value(ws, row_number, columns, "score_4_6"))),
        ("7-8 Strong", clean_text(cell_value(ws, row_number, columns, "score_7_8"))),
        ("9-10 Excellent", clean_text(cell_value(ws, row_number, columns, "score_9_10"))),
    ]

    parts: list[str] = []
    if measures:
        parts.append(f"What it measures: {measures}")
    scored_bands = [f"- {label}: {text}" for label, text in bands if text]
    if scored_bands:
        parts.append("Score bands:\n" + "\n".join(scored_bands))
    return "\n\n".join(parts)


def read_prompt_lookup(workbook_path: Path) -> dict[str, PromptInfo]:
    if not is_xlsx_workbook(workbook_path):
        return read_prompt_lookup_csv(workbook_path)

    wb = load_workbook(workbook_path, data_only=True)
    prompts: dict[str, PromptInfo] = {}

    for ws in wb.worksheets:
        if ws.title in RESULT_SHEET_NAMES or ws.title in GRADE_SHEET_NAMES:
            continue
        try:
            header_row, columns = find_header_row(
                ws,
                PROMPT_HEADER_ALIASES,
                {"test_id"},
            )
        except ValueError:
            continue

        for row_number in range(header_row + 1, ws.max_row + 1):
            test_id = clean_text(cell_value(ws, row_number, columns, "test_id"))
            if not test_id:
                continue
            info = PromptInfo(
                test_id=test_id,
                category=clean_text(cell_value(ws, row_number, columns, "category")),
                criterion=clean_text(cell_value(ws, row_number, columns, "criterion")),
                weight=parse_weight(cell_value(ws, row_number, columns, "weight")),
                eval_method=clean_text(cell_value(ws, row_number, columns, "eval_method")),
                prompt=clean_text(cell_value(ws, row_number, columns, "prompt")),
                input_material=clean_text(cell_value(ws, row_number, columns, "input_material")),
            )
            existing = prompts.get(test_id)
            if not existing or info.prompt or info.input_material:
                prompts[test_id] = info

    return prompts


def read_prompt_lookup_csv(csv_path: Path) -> dict[str, PromptInfo]:
    headers, rows = read_csv_dicts(csv_path)
    columns = find_csv_columns(headers, PROMPT_HEADER_ALIASES, {"test_id"})
    prompts: dict[str, PromptInfo] = {}

    for row in rows:
        test_id = clean_text(csv_value(row, columns, "test_id"))
        if not test_id:
            continue
        info = PromptInfo(
            test_id=test_id,
            category=clean_text(csv_value(row, columns, "category")),
            criterion=clean_text(csv_value(row, columns, "criterion")),
            weight=parse_weight_with_fallback(csv_value(row, columns, "weight"), 1.0),
            eval_method=clean_text(csv_value(row, columns, "eval_method")),
            prompt=clean_text(csv_value(row, columns, "prompt")),
            input_material=clean_text(csv_value(row, columns, "input_material")),
        )
        existing = prompts.get(test_id)
        if not existing or info.prompt or info.input_material:
            prompts[test_id] = info

    return prompts


def read_outputs(
    workbook_path: Path,
    sheet_name: str | None,
    only_tests: set[str],
    only_source_models: set[str],
    limit: int | None,
) -> tuple[list[TestOutput], list[SkippedOutput]]:
    if not is_xlsx_workbook(workbook_path):
        return read_outputs_csv(
            csv_path=workbook_path,
            only_tests=only_tests,
            only_source_models=only_source_models,
            limit=limit,
        )

    prompt_lookup = read_prompt_lookup(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = selected_sheet(wb, sheet_name, "Run Results")
    header_row, columns = find_header_row(
        ws,
        RESULT_HEADER_ALIASES,
        {"test_id", "source_model_key"},
    )

    outputs: list[TestOutput] = []
    skipped: list[SkippedOutput] = []

    for row_number in range(header_row + 1, ws.max_row + 1):
        test_id = clean_text(cell_value(ws, row_number, columns, "test_id"))
        source_model_key = clean_text(cell_value(ws, row_number, columns, "source_model_key"))
        if not test_id and not source_model_key:
            continue
        if only_tests and test_id not in only_tests:
            continue
        if only_source_models and source_model_key not in only_source_models:
            continue

        response = clean_text(cell_value(ws, row_number, columns, "response"))
        output_files = clean_text(cell_value(ws, row_number, columns, "output_files"))
        output_urls = clean_text(cell_value(ws, row_number, columns, "output_urls"))
        error = clean_text(cell_value(ws, row_number, columns, "error"))
        if error and not response:
            skipped.append(
                SkippedOutput(row_number, test_id, source_model_key, "source output errored")
            )
            continue
        if not response and not output_files and not output_urls:
            skipped.append(
                SkippedOutput(row_number, test_id, source_model_key, "missing source output")
            )
            continue

        prompt_info = prompt_lookup.get(test_id, PromptInfo(test_id=test_id))
        source_run_id = clean_text(cell_value(ws, row_number, columns, "source_run_id"))
        output_key_parts = [
            source_run_id or workbook_path.stem,
            str(row_number),
            source_model_key,
            test_id,
        ]
        outputs.append(
            TestOutput(
                row_number=row_number,
                output_key=":".join(output_key_parts),
                source_run_id=source_run_id,
                source_model_key=source_model_key,
                source_model_name=clean_text(
                    cell_value(ws, row_number, columns, "source_model_name")
                ),
                source_provider=clean_text(cell_value(ws, row_number, columns, "source_provider")),
                source_provider_model=clean_text(
                    cell_value(ws, row_number, columns, "source_provider_model")
                ),
                test_id=test_id,
                category=clean_text(cell_value(ws, row_number, columns, "category"))
                or prompt_info.category,
                criterion=clean_text(cell_value(ws, row_number, columns, "criterion"))
                or prompt_info.criterion,
                weight=parse_weight_with_fallback(
                    cell_value(ws, row_number, columns, "weight"),
                    prompt_info.weight,
                ),
                eval_method=clean_text(cell_value(ws, row_number, columns, "eval_method"))
                or prompt_info.eval_method,
                output_type=clean_text(cell_value(ws, row_number, columns, "output_type")),
                response=response,
                output_files=output_files,
                output_urls=output_urls,
                source_score=clean_text(cell_value(ws, row_number, columns, "source_score")),
                source_reasoning=clean_text(
                    cell_value(ws, row_number, columns, "source_reasoning")
                ),
                prompt=prompt_info.prompt,
                input_material=prompt_info.input_material,
                error=error,
            )
        )

    if limit is not None:
        skipped.extend(
            SkippedOutput(
                row.row_number,
                row.test_id,
                row.source_model_key,
                "outside --limit",
            )
            for row in outputs[limit:]
        )
        outputs = outputs[:limit]

    return outputs, skipped


def read_outputs_csv(
    csv_path: Path,
    only_tests: set[str],
    only_source_models: set[str],
    limit: int | None,
) -> tuple[list[TestOutput], list[SkippedOutput]]:
    headers, rows = read_csv_dicts(csv_path)
    columns = find_csv_columns(headers, RESULT_HEADER_ALIASES, {"test_id", "source_model_key"})
    prompt_lookup = read_prompt_lookup_csv(csv_path)

    outputs: list[TestOutput] = []
    skipped: list[SkippedOutput] = []

    for row_number, row in enumerate(rows, start=2):
        test_id = clean_text(csv_value(row, columns, "test_id"))
        source_model_key = clean_text(csv_value(row, columns, "source_model_key"))
        if not test_id and not source_model_key:
            continue
        if only_tests and test_id not in only_tests:
            continue
        if only_source_models and source_model_key not in only_source_models:
            continue

        response = clean_text(csv_value(row, columns, "response"))
        output_files = clean_text(csv_value(row, columns, "output_files"))
        output_urls = clean_text(csv_value(row, columns, "output_urls"))
        error = clean_text(csv_value(row, columns, "error"))
        if error and not response:
            skipped.append(
                SkippedOutput(row_number, test_id, source_model_key, "source output errored")
            )
            continue
        if not response and not output_files and not output_urls:
            skipped.append(
                SkippedOutput(row_number, test_id, source_model_key, "missing source output")
            )
            continue

        prompt_info = prompt_lookup.get(test_id, PromptInfo(test_id=test_id))
        source_run_id = clean_text(csv_value(row, columns, "source_run_id"))
        output_key_parts = [
            source_run_id or csv_path.stem,
            str(row_number),
            source_model_key,
            test_id,
        ]
        outputs.append(
            TestOutput(
                row_number=row_number,
                output_key=":".join(output_key_parts),
                source_run_id=source_run_id,
                source_model_key=source_model_key,
                source_model_name=clean_text(csv_value(row, columns, "source_model_name")),
                source_provider=clean_text(csv_value(row, columns, "source_provider")),
                source_provider_model=clean_text(
                    csv_value(row, columns, "source_provider_model")
                ),
                test_id=test_id,
                category=clean_text(csv_value(row, columns, "category")) or prompt_info.category,
                criterion=clean_text(csv_value(row, columns, "criterion")) or prompt_info.criterion,
                weight=parse_weight_with_fallback(
                    csv_value(row, columns, "weight"),
                    prompt_info.weight,
                ),
                eval_method=clean_text(csv_value(row, columns, "eval_method"))
                or prompt_info.eval_method,
                output_type=clean_text(csv_value(row, columns, "output_type")),
                response=response,
                output_files=output_files,
                output_urls=output_urls,
                source_score=clean_text(csv_value(row, columns, "source_score")),
                source_reasoning=clean_text(csv_value(row, columns, "source_reasoning")),
                prompt=prompt_info.prompt,
                input_material=prompt_info.input_material,
                error=error,
            )
        )

    if limit is not None:
        skipped.extend(
            SkippedOutput(
                row.row_number,
                row.test_id,
                row.source_model_key,
                "outside --limit",
            )
            for row in outputs[limit:]
        )
        outputs = outputs[:limit]

    return outputs, skipped


def read_rubric(workbook_path: Path, sheet_name: str | None) -> list[RubricEntry]:
    wb = load_workbook(workbook_path, data_only=True)
    try:
        ws, header_row, columns = rubric_sheet_and_header(wb, sheet_name)
    except ValueError as exc:
        raise ValueError(
            "The rubric workbook needs a sheet with a header row containing "
            "a Rubric, Scoring Rubric, Scoring Guidance, Grading Instructions, "
            "or What it measures column."
        ) from exc

    entries: list[RubricEntry] = []
    for row_number in range(header_row + 1, ws.max_row + 1):
        if not truthy(cell_value(ws, row_number, columns, "enabled"), default=True):
            continue
        rubric = composed_rubric_text(ws, row_number, columns)
        if not rubric:
            continue
        score_min = parse_score_bound(cell_value(ws, row_number, columns, "score_min"), 1.0)
        score_max = parse_score_bound(cell_value(ws, row_number, columns, "score_max"), 10.0)
        if score_max <= score_min:
            score_min, score_max = 1.0, 10.0
        entries.append(
            RubricEntry(
                row_number=row_number,
                test_id=clean_text(cell_value(ws, row_number, columns, "test_id")),
                category=clean_text(cell_value(ws, row_number, columns, "category")),
                criterion=clean_text(cell_value(ws, row_number, columns, "criterion")),
                website_field=clean_text(cell_value(ws, row_number, columns, "website_field")),
                weight=parse_weight(cell_value(ws, row_number, columns, "weight")),
                prompt=clean_text(cell_value(ws, row_number, columns, "prompt")),
                input_material=clean_text(cell_value(ws, row_number, columns, "input_material")),
                rubric=rubric,
                score_min=score_min,
                score_max=score_max,
            )
        )

    if not entries:
        raise ValueError("The rubric workbook did not contain any rubric rows.")
    return entries


def rubric_matches(output: TestOutput, entries: list[RubricEntry]) -> list[RubricEntry]:
    test_id = normalize_key(output.test_id)
    category = normalize_key(output.category)
    criterion = normalize_key(output.criterion)

    exact = [entry for entry in entries if normalize_key(entry.test_id) == test_id]
    if exact:
        return exact

    category_criterion = [
        entry
        for entry in entries
        if not entry.test_id
        and normalize_key(entry.category) == category
        and normalize_key(entry.criterion) == criterion
    ]
    if category_criterion:
        return category_criterion

    category_only = [
        entry
        for entry in entries
        if not entry.test_id
        and normalize_key(entry.category) == category
        and not normalize_key(entry.criterion)
    ]
    if category_only:
        return category_only

    return [
        entry
        for entry in entries
        if not entry.test_id and not entry.category and not entry.criterion
    ]


def rubric_context(output: TestOutput, entries: list[RubricEntry]) -> RubricContext:
    matches = rubric_matches(output, entries)
    if not matches:
        return RubricContext(
            rubric="",
            score_min=1.0,
            score_max=10.0,
            weight=output.weight,
            category=output.category,
            criterion=output.criterion,
            prompt=output.prompt,
            input_material=output.input_material,
            matched=False,
        )

    lines: list[str] = []
    for entry in matches:
        label_parts = []
        if entry.test_id:
            label_parts.append(f"Test ID {entry.test_id}")
        if entry.category:
            label_parts.append(entry.category)
        if entry.criterion:
            label_parts.append(entry.criterion)
        if entry.weight != 1.0:
            label_parts.append(f"weight {entry.weight:g}")
        label = " / ".join(label_parts) if label_parts else "General rubric"
        lines.append(f"{label}:\n{entry.rubric}")

    first = matches[0]
    prompt = output.prompt or first.prompt
    input_material = output.input_material or first.input_material
    return RubricContext(
        rubric="\n\n".join(lines),
        score_min=first.score_min,
        score_max=first.score_max,
        weight=first.weight if first.weight != 1.0 else output.weight,
        category=first.category or output.category,
        criterion=first.criterion or output.criterion,
        prompt=prompt,
        input_material=input_material,
        matched=True,
    )


def text_capable_models(config: dict[str, Any], models: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        model
        for model in models
        if "text" in model_capabilities(config, model) or "both" in model_capabilities(config, model)
    ]


def planned_pairs(
    models: list[dict[str, Any]],
    outputs: list[TestOutput],
    rubric_entries: list[RubricEntry],
    allow_missing_rubric: bool,
    include_self_judging: bool,
) -> tuple[list[tuple[dict[str, Any], TestOutput]], list[SkippedOutput]]:
    pairs: list[tuple[dict[str, Any], TestOutput]] = []
    skipped: list[SkippedOutput] = []
    for output in outputs:
        context = rubric_context(output, rubric_entries)
        if not context.matched and not allow_missing_rubric:
            skipped.append(
                SkippedOutput(
                    output.row_number,
                    output.test_id,
                    output.source_model_key,
                    "missing matching rubric",
                )
            )
            continue
        for model in models:
            if not include_self_judging and is_self_judge(model, output):
                continue
            pairs.append((model, output))
    return pairs, skipped


def grade_messages(
    output: TestOutput,
    context: RubricContext,
    score_only: bool = False,
) -> list[dict[str, str]]:
    prompt = context.prompt or "(not available in the source workbook)"
    input_material = context.input_material or "(none)"
    rubric = context.rubric or (
        "Use a strict 1-10 quality score for prompt adherence, correctness, "
        "completeness, and usefulness."
    )
    if score_only:
        response_shape = '{\n  "score": number\n}'
        response_rules = "- Return only the score JSON. Do not include explanation fields."
        missing_rule = (
            "- If the output cannot be evaluated from the available material, "
            "assign the lowest justified score."
        )
    else:
        response_shape = """{
  "score": number,
  "reasoning": "one or two concise sentences",
  "strengths": "brief note",
  "issues": "brief note"
}"""
        response_rules = "- Keep reasoning, strengths, and issues concise."
        missing_rule = (
            "- If the output cannot be evaluated from the available material, "
            "assign the lowest justified score and explain why."
        )

    user = f"""
Grade this AI model output using the rubric.

Test ID: {output.test_id}
Category: {output.category or "(none)"}
Criterion: {output.criterion or "(none)"}
Score range: {context.score_min:g} to {context.score_max:g}

Original prompt:
{prompt}

Input material:
{input_material}

Model under evaluation:
{output.source_model_name or output.source_model_key} ({output.source_provider_model or "unknown provider model"})

Model output:
{output.response}

Additional output files:
{output.output_files or "(none)"}

Additional output URLs:
{output.output_urls or "(none)"}

Rubric:
{rubric}

Return only JSON with:
{response_shape}

Rules:
- Judge only the model output above, not the model's reputation.
- Apply the rubric strictly and use the full score range when warranted.
- Penalize missing requested facts, unsupported claims, irrelevant content, and refusals when the prompt was answerable.
{missing_rule}
{response_rules}
""".strip()

    return [
        {
            "role": "system",
            "content": "You are a strict, consistent evaluator of AI prompt-test outputs.",
        },
        {"role": "user", "content": user},
    ]


def bounded_score(value: Any, score_min: float, score_max: float) -> float | None:
    try:
        score = float(value)
    except (TypeError, ValueError):
        return None
    return min(score_max, max(score_min, score))


def usage_reasoning_tokens(usage: dict[str, Any]) -> int | str:
    top_level = usage_token_count(usage, "reasoning_tokens")
    if top_level != "":
        return top_level
    return usage_detail_token_count(
        usage,
        "completion_tokens_details",
        "reasoning_tokens",
    )


def grade_output(
    config: dict[str, Any],
    grader: dict[str, Any],
    output: TestOutput,
    context: RubricContext,
    score_only: bool = False,
) -> tuple[float | None, str, str, str, dict[str, Any]]:
    grader_response, usage = call_openai_compatible(
        config,
        grader,
        grade_messages(output, context, score_only=score_only),
    )
    parsed = extract_json_object(grader_response)
    return (
        bounded_score(parsed.get("score"), context.score_min, context.score_max),
        clean_text(parsed.get("reasoning")),
        clean_text(parsed.get("strengths")),
        clean_text(parsed.get("issues")),
        usage,
    )


def result_row(
    grading_run_id: str,
    grader: dict[str, Any],
    output: TestOutput,
    context: RubricContext,
    score: float | None = None,
    reasoning: str = "",
    strengths: str = "",
    issues: str = "",
    latency_seconds: float | None = None,
    usage: dict[str, Any] | None = None,
    error: str = "",
) -> dict[str, Any]:
    usage_data = usage or {}
    return {
        "grading_run_id": grading_run_id,
        "timestamp": dt.datetime.now(dt.timezone.utc).isoformat(),
        "source_row": output.row_number,
        "source_run_id": output.source_run_id,
        "source_model_key": output.source_model_key,
        "source_model_name": output.source_model_name,
        "source_provider": output.source_provider,
        "source_provider_model": output.source_provider_model,
        "test_id": output.test_id,
        "category": context.category,
        "criterion": context.criterion,
        "weight": context.weight,
        "grader_model_key": grader.get("key", ""),
        "grader_model_name": grader.get("name", grader.get("key", "")),
        "grader_provider": grader.get("provider", ""),
        "grader_provider_model": grader.get("model", ""),
        "score": score if score is not None else "",
        "reasoning": reasoning,
        "strengths": strengths,
        "issues": issues,
        "rubric": context.rubric,
        "prompt": context.prompt,
        "input_material": context.input_material,
        "source_response": output.response,
        "source_output_files": output.output_files,
        "source_output_urls": output.output_urls,
        "source_score": output.source_score,
        "source_reasoning": output.source_reasoning,
        "latency_seconds": round(latency_seconds, 3) if latency_seconds is not None else "",
        "prompt_tokens": usage_token_count(usage_data, "prompt_tokens"),
        "completion_tokens": usage_token_count(usage_data, "completion_tokens"),
        "reasoning_tokens": usage_reasoning_tokens(usage_data),
        "total_tokens": usage_token_count(usage_data, "total_tokens"),
        "usage": usage_data,
        "error": error,
        "_output_key": output.output_key,
    }


def append_jsonl(path: Path, row: dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as handle:
        public_row = {key: value for key, value in row.items() if not key.startswith("_")}
        public_row["_output_key"] = row.get("_output_key", "")
        handle.write(json.dumps(public_row, ensure_ascii=False) + "\n")


def load_existing_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def make_output_dir(base: Path | None) -> Path:
    if base:
        base.mkdir(parents=True, exist_ok=True)
        return base

    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    output_dir = Path("outputs") / "prompt_grades" / stamp
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def default_website_seed_csv() -> Path | None:
    candidates = [
        Path.cwd() / "db/seeds/model_variants.csv",
        Path(__file__).resolve().parents[1] / "db/seeds/model_variants.csv",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def write_grades_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=GRADE_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    column: (
                        stringify_usage(row.get(column))
                        if column == "usage"
                        else row.get(column, "")
                    )
                    for column in GRADE_COLUMNS
                }
            )


def numeric_score(row: dict[str, Any]) -> float | None:
    value = row.get("score")
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def weighted_average_for_rows(rows: list[dict[str, Any]]) -> float | None:
    weighted_pairs = []
    for row in rows:
        score = numeric_score(row)
        if score is None:
            continue
        try:
            weight = float(row.get("weight") or 1.0)
        except (TypeError, ValueError):
            weight = 1.0
        weighted_pairs.append((score, weight))

    weight_total = sum(weight for _score, weight in weighted_pairs)
    if not weight_total:
        return None
    return sum(score * weight for score, weight in weighted_pairs) / weight_total


def summarize_grades(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_source: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_source.setdefault(clean_text(row.get("source_model_key")), []).append(row)

    summaries: list[dict[str, Any]] = []
    for source_model_key, source_rows in sorted(by_source.items()):
        scores = [score for row in source_rows if (score := numeric_score(row)) is not None]
        weighted_avg = weighted_average_for_rows(source_rows)
        latencies = [
            float(row["latency_seconds"])
            for row in source_rows
            if row.get("latency_seconds") not in (None, "")
        ]
        grader_keys = sorted(
            {
                clean_text(row.get("grader_model_key"))
                for row in source_rows
                if clean_text(row.get("grader_model_key"))
            }
        )
        outputs = sorted(
            {
                f"{clean_text(row.get('source_row'))}:{clean_text(row.get('test_id'))}"
                for row in source_rows
            }
        )
        summaries.append(
            {
                "source_model_key": source_model_key,
                "source_model_name": source_rows[0].get("source_model_name", ""),
                "source_provider": source_rows[0].get("source_provider", ""),
                "source_provider_model": source_rows[0].get("source_provider_model", ""),
                "outputs": len(outputs),
                "grade_rows": len(source_rows),
                "grader_models": ", ".join(grader_keys),
                "errors": sum(1 for row in source_rows if row.get("error")),
                "avg_score": average(scores),
                "weighted_avg": weighted_avg,
                "avg_latency_seconds": average(latencies),
            }
        )
    return summaries


def grader_keys_for_rows(rows: list[dict[str, Any]]) -> list[str]:
    return sorted(
        {
            clean_text(row.get("grader_model_key"))
            for row in rows
            if clean_text(row.get("grader_model_key"))
        }
    )


def summarize_model_scores(rows: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    grader_keys = grader_keys_for_rows(rows)
    by_source: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_source.setdefault(clean_text(row.get("source_model_key")), []).append(row)

    summaries: list[dict[str, Any]] = []
    for source_model_key, source_rows in sorted(by_source.items()):
        grader_averages: dict[str, float | None] = {}
        grader_weighted_averages: dict[str, float | None] = {}
        for grader_key in grader_keys:
            grader_rows = [
                row
                for row in source_rows
                if clean_text(row.get("grader_model_key")) == grader_key
            ]
            scores = [
                score
                for row in grader_rows
                if (score := numeric_score(row)) is not None
            ]
            grader_averages[grader_key] = average(scores)
            grader_weighted_averages[grader_key] = weighted_average_for_rows(grader_rows)

        available_grader_scores = [
            score for score in grader_averages.values() if score is not None
        ]
        available_weighted_scores = [
            score for score in grader_weighted_averages.values() if score is not None
        ]
        outputs = sorted(
            {
                f"{clean_text(row.get('source_row'))}:{clean_text(row.get('test_id'))}"
                for row in source_rows
            }
        )
        summaries.append(
            {
                "source_model_key": source_model_key,
                "source_model_name": source_rows[0].get("source_model_name", ""),
                "source_provider": source_rows[0].get("source_provider", ""),
                "source_provider_model": source_rows[0].get("source_provider_model", ""),
                "model_score": average(available_grader_scores),
                "weighted_model_score": average(available_weighted_scores),
                "outputs": len(outputs),
                "grade_rows": len(source_rows),
                "errors": sum(1 for row in source_rows if row.get("error")),
                "grader_averages": grader_averages,
            }
        )
    return grader_keys, summaries


def source_model_infos(
    outputs: Iterable[TestOutput],
    rows: Iterable[dict[str, Any]],
) -> list[dict[str, str]]:
    infos: dict[str, dict[str, str]] = {}

    def add_info(
        key: Any,
        name: Any = "",
        provider: Any = "",
        provider_model: Any = "",
    ) -> None:
        model_key = clean_text(key)
        if not model_key:
            return
        existing = infos.setdefault(
            model_key,
            {
                "key": model_key,
                "name": clean_text(name) or model_key,
                "provider": clean_text(provider),
                "provider_model": clean_text(provider_model),
            },
        )
        if clean_text(name):
            existing["name"] = clean_text(name)
        if clean_text(provider):
            existing["provider"] = clean_text(provider)
        if clean_text(provider_model):
            existing["provider_model"] = clean_text(provider_model)

    for output in outputs:
        add_info(
            output.source_model_key,
            output.source_model_name,
            output.source_provider,
            output.source_provider_model,
        )
    for row in rows:
        add_info(
            row.get("source_model_key"),
            row.get("source_model_name"),
            row.get("source_provider"),
            row.get("source_provider_model"),
        )
    return list(infos.values())


def rubric_test_infos(rubric_entries: list[RubricEntry], rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    tests: dict[str, dict[str, Any]] = {}
    for entry in rubric_entries:
        test_id = clean_text(entry.test_id)
        if not test_id or test_id in tests:
            continue
        tests[test_id] = {
            "test_id": test_id,
            "category": entry.category,
            "criterion": entry.criterion,
            "weight": entry.weight,
        }

    for row in rows:
        test_id = clean_text(row.get("test_id"))
        if not test_id or test_id in tests:
            continue
        try:
            weight = float(row.get("weight") or 1.0)
        except (TypeError, ValueError):
            weight = 1.0
        tests[test_id] = {
            "test_id": test_id,
            "category": clean_text(row.get("category")),
            "criterion": clean_text(row.get("criterion")),
            "weight": weight,
        }
    return list(tests.values())


def category_order(
    rubric_entries: list[RubricEntry],
    category_weights: dict[str, float],
    rows: Iterable[dict[str, Any]],
) -> list[str]:
    ordered: list[str] = []

    def add(category: Any) -> None:
        value = clean_text(category)
        if value and value not in ordered:
            ordered.append(value)

    for category in category_weights:
        add(category)
    for entry in rubric_entries:
        add(entry.category)
    if not ordered:
        for row in rows:
            add(row.get("category"))
    return ordered


def consensus_scores(
    rows: Iterable[dict[str, Any]],
    include_self_judging: bool = False,
) -> dict[tuple[str, str], float]:
    grouped: dict[tuple[str, str], list[float]] = {}
    for row in rows:
        if not include_self_judging and is_self_judge_row(row):
            continue
        score = numeric_score(row)
        if score is None:
            continue
        model_key = clean_text(row.get("source_model_key"))
        test_id = clean_text(row.get("test_id"))
        if not model_key or not test_id:
            continue
        grouped.setdefault((model_key, test_id), []).append(score)
    return {key: sum(scores) / len(scores) for key, scores in grouped.items() if scores}


def weighted_average_values(values: Iterable[tuple[float | None, float]]) -> float | None:
    pairs = [
        (score, weight)
        for score, weight in values
        if score is not None and weight not in (None, 0)
    ]
    if not pairs:
        return None
    weight_total = sum(weight for _score, weight in pairs)
    if not weight_total:
        return None
    return sum(score * weight for score, weight in pairs) / weight_total


def weighted_test_average(
    model_key: str,
    tests: Iterable[dict[str, Any]],
    consensus: dict[tuple[str, str], float],
) -> float | None:
    return weighted_average_values(
        (
            consensus.get((model_key, clean_text(test.get("test_id")))),
            float(test.get("weight") or 1.0),
        )
        for test in tests
    )


def weighted_results(
    model_infos: list[dict[str, str]],
    test_infos: list[dict[str, Any]],
    categories: list[str],
    category_weights: dict[str, float],
    consensus: dict[tuple[str, str], float],
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    tests_by_category = {
        category: [
            test
            for test in test_infos
            if normalize_key(clean_text(test.get("category"))) == normalize_key(category)
        ]
        for category in categories
    }

    for model in model_infos:
        model_key = model["key"]
        category_scores = {
            category: weighted_test_average(model_key, tests, consensus)
            for category, tests in tests_by_category.items()
        }
        overall = weighted_average_values(
            (
                score,
                category_weight_for(category, category_weights),
            )
            for category, score in category_scores.items()
        )
        results.append(
            {
                "model_key": model_key,
                "model_name": model["name"],
                "category_scores": category_scores,
                "overall": overall,
                "rank": "",
            }
        )

    for result in results:
        overall = result["overall"]
        if overall is None:
            continue
        result["rank"] = 1 + sum(
            1
            for other in results
            if other["overall"] is not None and other["overall"] > overall
        )
    return results


def rounded_score(value: float | None) -> float | str:
    return round(value, 4) if value is not None else ""


def text_preview(value: Any, limit: int = 500) -> str:
    text = clean_text(value)
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "..."


def populate_grade_sheets(
    wb: Any,
    rows: list[dict[str, Any]],
    skipped: list[SkippedOutput],
    rubric_entries: list[RubricEntry] | None = None,
    category_weights: dict[str, float] | None = None,
    outputs: list[TestOutput] | None = None,
    include_self_judging: bool = False,
) -> None:
    rubric_entries = rubric_entries or []
    category_weights = category_weights or {}
    outputs = outputs or []

    summary_ws = wb.create_sheet("Grade Summary", 0)
    model_scores_ws = wb.create_sheet("Model Scores", 1)
    results_ws = wb.create_sheet("Results", 2)
    consensus_ws = wb.create_sheet("Consensus", 3)
    ai_scores_ws = wb.create_sheet("AI Scores", 4)
    grades_ws = wb.create_sheet("Grades", 5)
    skipped_ws = wb.create_sheet("Skipped Outputs", 6)

    summary_ws.append(
        [
            "Source Model Key",
            "Source Model Name",
            "Source Provider",
            "Source Provider Model",
            "Outputs",
            "Grade Rows",
            "Grader Models",
            "Errors",
            "Average Score",
            "Weighted Average",
            "Average Latency Seconds",
        ]
    )
    for summary in summarize_grades(rows):
        summary_ws.append(
            [
                summary["source_model_key"],
                summary["source_model_name"],
                summary["source_provider"],
                summary["source_provider_model"],
                summary["outputs"],
                summary["grade_rows"],
                summary["grader_models"],
                summary["errors"],
                summary["avg_score"],
                summary["weighted_avg"],
                summary["avg_latency_seconds"],
            ]
        )

    grader_keys, model_score_rows = summarize_model_scores(rows)
    model_scores_ws.append(
        [
            "Source Model Key",
            "Source Model Name",
            "Source Provider",
            "Source Provider Model",
            "Model Score",
            "Weighted Model Score",
            "Outputs",
            "Grade Rows",
            "Errors",
            *[f"{grader_key} Average" for grader_key in grader_keys],
        ]
    )
    for summary in model_score_rows:
        model_scores_ws.append(
            [
                summary["source_model_key"],
                summary["source_model_name"],
                summary["source_provider"],
                summary["source_provider_model"],
                summary["model_score"],
                summary["weighted_model_score"],
                summary["outputs"],
                summary["grade_rows"],
                summary["errors"],
                *[
                    summary["grader_averages"].get(grader_key)
                    for grader_key in grader_keys
                ],
            ]
        )

    model_infos = source_model_infos(outputs, rows)
    test_infos = rubric_test_infos(rubric_entries, rows)
    categories = category_order(rubric_entries, category_weights, rows)
    consensus = consensus_scores(rows, include_self_judging=include_self_judging)
    result_rows = weighted_results(
        model_infos=model_infos,
        test_infos=test_infos,
        categories=categories,
        category_weights=category_weights,
        consensus=consensus,
    )

    results_ws.append(
        [
            "Model Key",
            "Model",
            *categories,
            "Overall",
            "Rank",
        ]
    )
    for result in sorted(
        result_rows,
        key=lambda row: (
            row["rank"] if row["rank"] != "" else 10_000,
            row["model_name"],
        ),
    ):
        results_ws.append(
            [
                result["model_key"],
                result["model_name"],
                *[
                    rounded_score(result["category_scores"].get(category))
                    for category in categories
                ],
                rounded_score(result["overall"]),
                result["rank"],
            ]
        )

    consensus_ws.append(
        [
            "Test ID",
            "Category",
            "Criterion",
            "Weight",
            *[model["name"] for model in model_infos],
        ]
    )
    for test in test_infos:
        consensus_ws.append(
            [
                test["test_id"],
                test["category"],
                test["criterion"],
                test["weight"],
                *[
                    rounded_score(consensus.get((model["key"], test["test_id"])))
                    for model in model_infos
                ],
            ]
        )

    ai_scores_ws.append(
        [
            "Test ID",
            "Model Evaluated",
            "Judge (AI)",
            "Output (auto)",
            "Score (1-10)",
            "Self-judge?",
            "Rationale",
            "Date",
        ]
    )
    for row in rows:
        ai_scores_ws.append(
            [
                row.get("test_id", ""),
                row.get("source_model_name") or row.get("source_model_key", ""),
                row.get("grader_model_name") or row.get("grader_model_key", ""),
                text_preview(row.get("source_response")),
                row.get("score", ""),
                "yes" if is_self_judge_row(row) else "",
                row.get("reasoning", ""),
                clean_text(row.get("timestamp"))[:10],
            ]
        )

    grades_ws.append(GRADE_COLUMNS)
    for row in rows:
        grades_ws.append(
            [
                stringify_usage(row.get(column)) if column == "usage" else row.get(column, "")
                for column in GRADE_COLUMNS
            ]
        )

    skipped_ws.append(["Source Row", "Test ID", "Source Model Key", "Reason"])
    for skipped_output in skipped:
        skipped_ws.append(
            [
                skipped_output.row_number,
                skipped_output.test_id,
                skipped_output.source_model_key,
                skipped_output.reason,
            ]
        )

    for ws in [
        summary_ws,
        model_scores_ws,
        results_ws,
        consensus_ws,
        ai_scores_ws,
        grades_ws,
        skipped_ws,
    ]:
        style_header(ws)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    set_widths(
        summary_ws,
        {1: 18, 2: 28, 3: 18, 4: 34, 5: 10, 6: 12, 7: 34, 8: 10, 9: 14, 10: 16, 11: 22},
    )
    set_widths(
        model_scores_ws,
        {
            1: 18,
            2: 28,
            3: 18,
            4: 34,
            5: 14,
            6: 20,
            7: 10,
            8: 12,
            9: 10,
            **{index: 18 for index in range(10, 10 + len(grader_keys))},
        },
    )
    set_widths(
        results_ws,
        {
            1: 18,
            2: 28,
            **{index: 18 for index in range(3, 3 + len(categories))},
            3 + len(categories): 12,
            4 + len(categories): 10,
        },
    )
    set_widths(
        consensus_ws,
        {
            1: 12,
            2: 24,
            3: 28,
            4: 10,
            **{index: 18 for index in range(5, 5 + len(model_infos))},
        },
    )
    set_widths(
        ai_scores_ws,
        {1: 12, 2: 28, 3: 28, 4: 80, 5: 12, 6: 12, 7: 48, 8: 14},
    )
    set_widths(
        grades_ws,
        {
            1: 22,
            2: 22,
            3: 10,
            4: 20,
            5: 18,
            6: 26,
            7: 18,
            8: 34,
            9: 12,
            10: 24,
            11: 28,
            12: 10,
            13: 18,
            14: 28,
            15: 16,
            16: 34,
            17: 10,
            18: 48,
            19: 38,
            20: 38,
            21: 64,
            22: 64,
            23: 48,
            24: 80,
            25: 42,
            26: 42,
            27: 12,
            28: 42,
            29: 16,
            30: 14,
            31: 17,
            32: 17,
            33: 14,
            34: 42,
            35: 48,
        },
    )
    set_widths(skipped_ws, {1: 12, 2: 14, 3: 22, 4: 42})


def write_live_grades_workbook(
    output_path: Path,
    rows: list[dict[str, Any]],
    skipped: list[SkippedOutput],
    rubric_entries: list[RubricEntry],
    category_weights: dict[str, float],
    outputs: list[TestOutput],
    include_self_judging: bool,
) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    populate_grade_sheets(
        wb,
        rows,
        skipped,
        rubric_entries=rubric_entries,
        category_weights=category_weights,
        outputs=outputs,
        include_self_judging=include_self_judging,
    )
    save_workbook_atomic(wb, output_path)


def write_augmented_grades_workbook(
    source_workbook: Path,
    output_path: Path,
    rows: list[dict[str, Any]],
    skipped: list[SkippedOutput],
    rubric_entries: list[RubricEntry],
    category_weights: dict[str, float],
    outputs: list[TestOutput],
    include_self_judging: bool,
) -> None:
    wb = load_workbook(source_workbook)
    for sheet_name in GRADE_SHEET_NAMES:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    populate_grade_sheets(
        wb,
        rows,
        skipped,
        rubric_entries=rubric_entries,
        category_weights=category_weights,
        outputs=outputs,
        include_self_judging=include_self_judging,
    )
    save_workbook_atomic(wb, output_path)


def read_csv_dicts(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), [dict(row) for row in reader]


def write_csv_dicts(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def normalized_name(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_text(value).lower())


def website_field_tests(headers: list[str], rubric_entries: list[RubricEntry]) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {}
    for source in (WEBSITE_DIRECT_FIELD_TEST_IDS, WEBSITE_COMPOSITE_FIELD_TEST_IDS):
        for field, test_ids in source.items():
            if field in headers:
                mapping[field] = list(test_ids)
    for entry in rubric_entries:
        field = clean_text(entry.website_field)
        test_id = clean_text(entry.test_id)
        if not field or not test_id or field not in headers:
            continue
        mapping.setdefault(field, [])
        if test_id not in mapping[field]:
            mapping[field].append(test_id)
    return mapping


def website_score_fields(headers: list[str]) -> list[str]:
    return [
        header
        for header in headers
        if header.startswith("score_") or header.endswith("_score")
    ]


def weighted_consensus_for_tests(
    model_key: str,
    test_ids: list[str],
    test_info_by_id: dict[str, dict[str, Any]],
    consensus: dict[tuple[str, str], float],
) -> tuple[float | None, list[str]]:
    values: list[tuple[float | None, float]] = []
    used_tests: list[str] = []
    for test_id in test_ids:
        score = consensus.get((model_key, test_id))
        if score is None:
            continue
        weight = float(test_info_by_id.get(test_id, {}).get("weight") or 1.0)
        values.append((score, weight))
        used_tests.append(test_id)
    return weighted_average_values(values), used_tests


def write_website_upload_files(
    seed_csv_path: Path,
    upload_csv_path: Path,
    audit_csv_path: Path,
    rows: list[dict[str, Any]],
    rubric_entries: list[RubricEntry],
    outputs: list[TestOutput],
    include_self_judging: bool,
    skipped: list[SkippedOutput] | None = None,
) -> tuple[int, int]:
    headers, seed_rows = read_csv_dicts(seed_csv_path)
    output_rows = [dict(row) for row in seed_rows]
    model_infos = source_model_infos(outputs, rows)
    test_infos = rubric_test_infos(rubric_entries, rows)
    test_info_by_id = {clean_text(test["test_id"]): test for test in test_infos}
    consensus = consensus_scores(rows, include_self_judging=include_self_judging)
    field_tests = website_field_tests(headers, rubric_entries)
    score_fields = website_score_fields(headers)
    preserved_source_keys = {
        clean_text(skipped_output.source_model_key)
        for skipped_output in (skipped or [])
        if clean_text(skipped_output.source_model_key)
        and skipped_output.reason == "source model already scored"
    }
    skipped_source_keys = {
        clean_text(skipped_output.source_model_key)
        for skipped_output in (skipped or [])
        if clean_text(skipped_output.source_model_key)
        and skipped_output.reason in SCORE_CLEARING_SKIPPED_REASONS
    } - preserved_source_keys

    sources_by_key = {model["key"]: model for model in model_infos}
    sources_by_name: dict[str, list[dict[str, str]]] = {}
    for model in model_infos:
        sources_by_name.setdefault(normalized_name(model["name"]), []).append(model)

    matched = 0
    updated = 0
    audit_rows: list[dict[str, Any]] = []

    for seed_row in output_rows:
        model_id = clean_text(seed_row.get("model_id_string"))
        source = sources_by_key.get(model_id)
        if not source:
            name_matches = sources_by_name.get(normalized_name(seed_row.get("name")), [])
            if len(name_matches) == 1:
                source = name_matches[0]
                if not model_id:
                    seed_row["model_id_string"] = source["key"]
        if not source:
            if model_id and model_id in skipped_source_keys:
                for field in score_fields:
                    seed_row[field] = ""
                updated += 1
            continue

        matched += 1
        row_updated = False
        for field, test_ids in field_tests.items():
            value, used_tests = weighted_consensus_for_tests(
                source["key"],
                test_ids,
                test_info_by_id,
                consensus,
            )
            if value is None:
                continue
            previous_value = seed_row.get(field, "")
            db_value = str(round_half_up(value))
            seed_row[field] = db_value
            row_updated = True
            audit_rows.append(
                {
                    "source_model_key": source["key"],
                    "source_model_name": source["name"],
                    "tool_name": seed_row.get("tool_name", ""),
                    "variant_name": seed_row.get("name", ""),
                    "db_field": field,
                    "mapped_test_ids": ", ".join(test_ids),
                    "used_test_ids": ", ".join(used_tests),
                    "raw_weighted_average": round(value, 4),
                    "db_value": db_value,
                    "previous_value": previous_value,
                }
            )
        if row_updated:
            updated += 1

    write_csv_dicts(upload_csv_path, headers, output_rows)
    write_csv_dicts(
        audit_csv_path,
        [
            "source_model_key",
            "source_model_name",
            "tool_name",
            "variant_name",
            "db_field",
            "mapped_test_ids",
            "used_test_ids",
            "raw_weighted_average",
            "db_value",
            "previous_value",
        ],
        audit_rows,
    )
    return matched, updated


def selected_output_keys(outputs: Iterable[TestOutput]) -> dict[str, int]:
    return {output.output_key: index for index, output in enumerate(outputs)}


def completed_grade_keys(
    existing_rows: Iterable[dict[str, Any]],
    selected_keys: set[tuple[str, str]],
) -> set[tuple[str, str]]:
    completed: set[tuple[str, str]] = set()
    for row in existing_rows:
        key = (clean_text(row.get("grader_model_key")), clean_text(row.get("_output_key")))
        if key in selected_keys and (row.get("score") not in (None, "") or row.get("error")):
            completed.add(key)
    return completed


def filter_scored_source_outputs(
    outputs: list[TestOutput],
    scored_model_keys: dict[str, list[str]],
) -> tuple[list[TestOutput], list[SkippedOutput]]:
    selected: list[TestOutput] = []
    skipped: list[SkippedOutput] = []
    for output in outputs:
        source_model_key = clean_text(output.source_model_key)
        if source_model_key and source_model_key in scored_model_keys:
            skipped.append(
                SkippedOutput(
                    output.row_number,
                    output.test_id,
                    output.source_model_key,
                    "source model already scored",
                )
            )
        else:
            selected.append(output)
    return selected, skipped


def print_plan(
    config: dict[str, Any],
    models: list[dict[str, Any]],
    outputs: list[TestOutput],
    skipped: list[SkippedOutput],
    rubric_entries: list[RubricEntry],
    category_weights: dict[str, float],
    pairs: list[tuple[dict[str, Any], TestOutput]],
    parallel_products: bool,
    requested_workers: int,
    include_self_judging: bool,
) -> None:
    print(f"Grading models enabled: {len(models)}")
    for model in models:
        capabilities = ", ".join(sorted(model_capabilities(config, model)))
        print(
            f"  - {model.get('key')}: {model.get('name')} "
            f"({model.get('model')}; {capabilities})"
        )

    print(f"Source outputs selected: {len(outputs)}")
    by_source: dict[str, int] = {}
    for output in outputs:
        by_source[output.source_model_key] = by_source.get(output.source_model_key, 0) + 1
    for source, count in sorted(by_source.items()):
        print(f"  - {source}: {count}")

    print(f"Rubric rows: {len(rubric_entries)}")
    print(f"Category weights: {len(category_weights)}")
    for category, weight in category_weights.items():
        print(f"  - {category}: {weight:g}")
    if not include_self_judging:
        planned_without_self = len(pairs)
        planned_with_self = 0
        for output in outputs:
            if not rubric_context(output, rubric_entries).matched:
                continue
            planned_with_self += len(models)
        skipped_self = planned_with_self - planned_without_self
        print(f"Self-judging calls excluded: {max(0, skipped_self)}")
    missing_rubric = [
        output for output in outputs if not rubric_context(output, rubric_entries).matched
    ]
    print(f"Outputs without matching rubric: {len(missing_rubric)}")
    print(f"Skipped outputs: {len(skipped)}")
    if skipped:
        print("Skipped output reasons:")
        reason_counts = Counter(skipped_output.reason for skipped_output in skipped)
        for reason, count in reason_counts.most_common():
            print(f"  - {reason}: {count}")
    for skipped_output in skipped[:20]:
        print(
            f"  - row {skipped_output.row_number} "
            f"{skipped_output.source_model_key}/{skipped_output.test_id}: "
            f"{skipped_output.reason}"
        )
    if len(skipped) > 20:
        print(f"  - ... {len(skipped) - 20} more")

    print(f"Total grading API calls: {len(pairs)}")
    groups = group_pairs_by_product(config, pairs)
    workers = product_worker_count(parallel_products, requested_workers, len(groups))
    mode = "parallel" if parallel_products and workers > 1 else "serial"
    print(f"Product lanes: {len(groups)} ({mode}; workers: {workers})")
    for lane, lane_pairs in groups.items():
        model_keys = sorted({clean_text(model.get("key")) for model, _output in lane_pairs})
        print(
            f"  - {product_lane_label(lane)}: "
            f"{len(model_keys)} grader model(s), {len(lane_pairs)} call(s)"
        )


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-workbook",
        required=True,
        help="Path to a testing-app output workbook, usually model_test_results.xlsx or responses.xlsx.",
    )
    parser.add_argument(
        "--models-workbook",
        required=True,
        help="Path to an .xlsx grader model list. Same format as the model testing app.",
    )
    parser.add_argument(
        "--rubric-workbook",
        required=True,
        help="Path to an .xlsx rubric workbook with a Rubric or Scoring Guidance column.",
    )
    parser.add_argument("--results-sheet", help="Results sheet name. Defaults to Run Results.")
    parser.add_argument("--rubric-sheet", help="Rubric sheet name. Defaults to the first sheet.")
    parser.add_argument("--output-dir", help="Output directory. Defaults to outputs/prompt_grades/<timestamp>.")
    parser.add_argument("--dry-run", action="store_true", help="Print the grading plan without API calls.")
    parser.add_argument("--force", action="store_true", help="Ignore existing grades.jsonl in output-dir.")
    parser.add_argument(
        "--max-tokens",
        type=int,
        default=DEFAULT_MAX_TOKENS,
        help="Maximum judge response tokens. Defaults to 800.",
    )
    parser.add_argument(
        "--excel-every",
        type=int,
        default=1,
        help="Refresh grades.xlsx every N completed rows. Use 0 to disable live Excel output.",
    )
    parser.add_argument("--limit", type=int, help="Limit the number of selected source outputs.")
    parser.add_argument("--only-tests", help="Comma-separated Test IDs to grade.")
    parser.add_argument("--only-source-models", help="Comma-separated source model keys to grade.")
    parser.add_argument("--only-models", help="Comma-separated grader model keys from the model workbook.")
    parser.add_argument(
        "--first-grader-only",
        action="store_true",
        help="Use only the first enabled text-capable grader model after filters.",
    )
    parser.add_argument(
        "--score-only",
        action="store_true",
        help="Ask graders to return only a numeric score JSON object.",
    )
    parser.add_argument(
        "--allow-missing-rubric",
        action="store_true",
        help="Grade rows without a matching rubric using generic guidance instead of skipping them.",
    )
    parser.add_argument(
        "--include-self-judging",
        action="store_true",
        help="Allow a grader model to score its own output. Defaults to excluding self-judging.",
    )
    parser.add_argument(
        "--website-seed-csv",
        help=(
            "Existing website db/seeds/model_variants.csv to use as metadata. "
            "Defaults to db/seeds/model_variants.csv when the script can find it."
        ),
    )
    parser.add_argument(
        "--website-upload-csv",
        help=(
            "Where to write the website-ready model_variants CSV. Defaults to "
            "<output-dir>/db_upload/model_variants_db_upload.csv when a seed CSV is available."
        ),
    )
    parser.add_argument(
        "--update-website-seed",
        action="store_true",
        help="Overwrite --website-seed-csv with the generated website upload CSV after grading.",
    )
    parser.add_argument(
        "--skip-scored-source-models",
        action="store_true",
        help="Skip source model outputs whose model_id_string already has website scores.",
    )
    parser.add_argument(
        "--rate-limit-skip-after",
        type=int,
        default=3,
        help="Skip remaining rows for a grader model after this many rate-limit errors.",
    )
    parser.add_argument(
        "--parallel-products",
        action="store_true",
        help="Run different provider/product lanes concurrently.",
    )
    parser.add_argument(
        "--product-workers",
        type=int,
        default=0,
        help="Maximum concurrent product lanes when --parallel-products is set.",
    )
    parser.add_argument("--verbose-errors", action="store_true", help="Print tracebacks for failed calls.")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    if args.max_tokens < 1:
        raise ValueError("--max-tokens must be greater than 0.")
    if args.excel_every < 0:
        raise ValueError("--excel-every must be 0 or greater.")
    if args.limit is not None and args.limit < 1:
        raise ValueError("--limit must be greater than 0.")
    if args.rate_limit_skip_after < 0:
        raise ValueError("--rate-limit-skip-after must be 0 or greater.")
    if args.product_workers < 0:
        raise ValueError("--product-workers must be 0 or greater.")

    results_workbook = Path(args.results_workbook).expanduser()
    models_workbook = Path(args.models_workbook).expanduser()
    rubric_workbook = Path(args.rubric_workbook).expanduser()
    requested_output_dir = Path(args.output_dir).expanduser() if args.output_dir else None
    requested_website_seed_csv = (
        Path(args.website_seed_csv).expanduser() if args.website_seed_csv else None
    )
    requested_website_upload_csv = (
        Path(args.website_upload_csv).expanduser() if args.website_upload_csv else None
    )

    config = read_model_workbook(models_workbook, default_config())
    config.setdefault("request", {})["max_tokens"] = args.max_tokens
    models = text_capable_models(config, enabled_models(config, args.only_models))
    if args.first_grader_only and len(models) > 1:
        models = models[:1]
    if not models:
        raise ValueError("No enabled text-capable grader models found.")

    website_seed_csv = requested_website_seed_csv or default_website_seed_csv()

    only_tests = split_filter(args.only_tests)
    only_source_models = split_filter(args.only_source_models)
    outputs, skipped = read_outputs(
        workbook_path=results_workbook,
        sheet_name=args.results_sheet,
        only_tests=only_tests,
        only_source_models=only_source_models,
        limit=args.limit,
    )
    source_outputs_before_score_skip = len(outputs)
    scored_skipped: list[SkippedOutput] = []
    if args.skip_scored_source_models:
        if not website_seed_csv:
            raise ValueError("--skip-scored-source-models requires --website-seed-csv.")
        if not website_seed_csv.exists():
            raise ValueError(f"Website seed CSV not found: {website_seed_csv}")
        outputs, scored_skipped = filter_scored_source_outputs(
            outputs,
            read_scored_model_keys(website_seed_csv),
        )
        skipped.extend(scored_skipped)
    rubric_entries = read_rubric(rubric_workbook, args.rubric_sheet)
    category_weights = read_category_weights(rubric_workbook)
    pairs, rubric_skipped = planned_pairs(
        models=models,
        outputs=outputs,
        rubric_entries=rubric_entries,
        allow_missing_rubric=args.allow_missing_rubric,
        include_self_judging=args.include_self_judging,
    )
    skipped.extend(rubric_skipped)

    print_plan(
        config=config,
        models=models,
        outputs=outputs,
        skipped=skipped,
        rubric_entries=rubric_entries,
        category_weights=category_weights,
        pairs=pairs,
        parallel_products=args.parallel_products,
        requested_workers=args.product_workers,
        include_self_judging=args.include_self_judging,
    )

    if args.dry_run:
        return 0
    if not outputs:
        if (
            args.skip_scored_source_models
            and source_outputs_before_score_skip > 0
            and len(scored_skipped) == source_outputs_before_score_skip
        ):
            print(
                "No unscored source outputs selected. All valid source outputs "
                "already have website scores, so there is nothing to grade."
            )
            return 0
        raise ValueError("No source outputs selected.")
    if not pairs:
        raise ValueError("No grading calls planned. Check rubric coverage and filters.")

    validate_api_keys(config, pairs)
    validate_openrouter_model_ids(config, pairs)

    output_dir = make_output_dir(requested_output_dir)
    grading_run_id = output_dir.name
    jsonl_path = output_dir / "grades.jsonl"
    live_xlsx_path = output_dir / "grades.xlsx"
    existing_rows = [] if args.force else load_existing_jsonl(jsonl_path)

    selected_keys = {
        (clean_text(model.get("key")), output.output_key)
        for model, output in pairs
    }
    completed = completed_grade_keys(existing_rows, selected_keys)
    all_rows = list(existing_rows)
    output_order = selected_output_keys(outputs)
    pair_order = {
        (clean_text(model.get("key")), output.output_key): index
        for index, (model, output) in enumerate(pairs)
    }

    def ordered_rows() -> list[dict[str, Any]]:
        return sorted(
            all_rows,
            key=lambda row: (
                pair_order.get(
                    (clean_text(row.get("grader_model_key")), clean_text(row.get("_output_key"))),
                    len(pair_order),
                ),
                output_order.get(clean_text(row.get("_output_key")), len(output_order)),
                clean_text(row.get("timestamp")),
            ),
        )

    if args.excel_every:
        write_live_grades_workbook(
            live_xlsx_path,
            ordered_rows(),
            skipped,
            rubric_entries,
            category_weights,
            outputs,
            args.include_self_judging,
        )

    total = len(pairs)
    done_count = len(completed)
    progress_lock = threading.Lock()
    output_lock = threading.Lock()

    def next_progress() -> int:
        nonlocal done_count
        with progress_lock:
            done_count += 1
            return done_count

    def persist_row(row: dict[str, Any]) -> None:
        with output_lock:
            append_jsonl(jsonl_path, row)
            all_rows.append(row)
            if args.excel_every and len(all_rows) % args.excel_every == 0:
                write_live_grades_workbook(
                    live_xlsx_path,
                    ordered_rows(),
                    skipped,
                    rubric_entries,
                    category_weights,
                    outputs,
                    args.include_self_judging,
                )

    def run_pair(grader: dict[str, Any], output: TestOutput) -> dict[str, Any]:
        context = rubric_context(output, rubric_entries)
        started = time.monotonic()
        score, reasoning, strengths, issues, usage = grade_output(
            config=config,
            grader=grader,
            output=output,
            context=context,
            score_only=args.score_only,
        )
        return result_row(
            grading_run_id=grading_run_id,
            grader=grader,
            output=output,
            context=context,
            score=score,
            reasoning=reasoning,
            strengths=strengths,
            issues=issues,
            latency_seconds=time.monotonic() - started,
            usage=usage,
        )

    product_groups = group_pairs_by_product(config, pairs)
    workers = product_worker_count(args.parallel_products, args.product_workers, len(product_groups))

    def run_lane(lane: str, lane_pairs: list[tuple[dict[str, Any], TestOutput]]) -> None:
        rate_limit_errors_by_model: dict[str, int] = {}
        rate_limited_models: set[str] = set()
        lane_label = product_lane_label(lane)
        show_lane = args.parallel_products and len(product_groups) > 1

        for grader, output in lane_pairs:
            grader_key = clean_text(grader.get("key"))
            key = (grader_key, output.output_key)
            if key in completed:
                print(f"Skipping existing {grader_key} / row {output.row_number}")
                continue
            if grader_key in rate_limited_models:
                print(f"Skipping rate-limited {grader_key} / row {output.row_number}")
                continue

            current = next_progress()
            lane_prefix = f"[{lane_label}] " if show_lane else ""
            print(
                f"[{current}/{total}] {lane_prefix}{grader_key} grades "
                f"{output.source_model_key} / {output.test_id}",
                flush=True,
            )
            started = time.monotonic()
            context = rubric_context(output, rubric_entries)
            try:
                row = run_pair(grader, output)
            except Exception as exc:  # pragma: no cover - depends on remote APIs
                if isinstance(exc, RateLimitError):
                    rate_limit_errors_by_model[grader_key] = (
                        rate_limit_errors_by_model.get(grader_key, 0) + 1
                    )
                    if (
                        args.rate_limit_skip_after
                        and rate_limit_errors_by_model[grader_key] >= args.rate_limit_skip_after
                    ):
                        rate_limited_models.add(grader_key)
                row = result_row(
                    grading_run_id=grading_run_id,
                    grader=grader,
                    output=output,
                    context=context,
                    latency_seconds=time.monotonic() - started,
                    error=f"{type(exc).__name__}: {exc}",
                )
                if args.verbose_errors:
                    print(traceback.format_exc(), file=sys.stderr)
                else:
                    print(f"  error: {row['error']}", file=sys.stderr)
                if grader_key in rate_limited_models:
                    print(
                        "  rate limit: skipping remaining rows for "
                        f"{grader_key} after {args.rate_limit_skip_after} rate-limit errors",
                        file=sys.stderr,
                    )

            persist_row(row)

    if workers > 1:
        print(f"Running product lanes in parallel with {workers} workers.", flush=True)
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
            future_to_lane = {
                executor.submit(run_lane, lane, lane_pairs): lane
                for lane, lane_pairs in product_groups.items()
            }
            for future in concurrent.futures.as_completed(future_to_lane):
                lane = future_to_lane[future]
                try:
                    future.result()
                except Exception as exc:
                    raise RuntimeError(f"Product lane {product_lane_label(lane)} failed.") from exc
    else:
        for lane, lane_pairs in product_groups.items():
            run_lane(lane, lane_pairs)

    final_rows = ordered_rows()
    if args.excel_every:
        write_live_grades_workbook(
            live_xlsx_path,
            final_rows,
            skipped,
            rubric_entries,
            category_weights,
            outputs,
            args.include_self_judging,
        )
    write_grades_csv(output_dir / "grades.csv", final_rows)
    if is_xlsx_workbook(results_workbook):
        write_augmented_grades_workbook(
            source_workbook=results_workbook,
            output_path=output_dir / "prompt_output_grades.xlsx",
            rows=final_rows,
            skipped=skipped,
            rubric_entries=rubric_entries,
            category_weights=category_weights,
            outputs=outputs,
            include_self_judging=args.include_self_judging,
        )
    else:
        write_live_grades_workbook(
            output_dir / "prompt_output_grades.xlsx",
            final_rows,
            skipped,
            rubric_entries,
            category_weights,
            outputs,
            args.include_self_judging,
        )

    website_upload_csv = requested_website_upload_csv
    if website_seed_csv:
        if not website_seed_csv.exists():
            raise ValueError(f"Website seed CSV not found: {website_seed_csv}")
        if not website_upload_csv:
            website_upload_csv = output_dir / "db_upload" / "model_variants_db_upload.csv"
        website_audit_csv = website_upload_csv.parent / "model_variant_score_audit.csv"
        matched, updated = write_website_upload_files(
            seed_csv_path=website_seed_csv,
            upload_csv_path=website_upload_csv,
            audit_csv_path=website_audit_csv,
            rows=final_rows,
            rubric_entries=rubric_entries,
            outputs=outputs,
            include_self_judging=args.include_self_judging,
            skipped=skipped,
        )
        print(f"Wrote {website_upload_csv}")
        print(f"Wrote {website_audit_csv}")
        print(f"Website seed rows matched: {matched}; updated: {updated}")
        if args.update_website_seed:
            website_seed_csv.write_text(website_upload_csv.read_text(encoding="utf-8"), encoding="utf-8")
            print(f"Updated {website_seed_csv}")
    elif args.update_website_seed or requested_website_upload_csv:
        raise ValueError("Provide --website-seed-csv to generate or update a website seed CSV.")

    print(f"Wrote {jsonl_path}")
    print(f"Wrote {live_xlsx_path}")
    print(f"Wrote {output_dir / 'grades.csv'}")
    print(f"Wrote {output_dir / 'prompt_output_grades.xlsx'}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv[1:]))
    except KeyboardInterrupt:
        print("Cancelled.", file=sys.stderr)
        raise SystemExit(130)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1)
