#!/usr/bin/env python3
"""Run the model prompt-test workbook against configured chat models.

The runner intentionally uses a conservative, portable API surface:
OpenAI-compatible /chat/completions endpoints via urllib, plus openpyxl for
reading and writing the workbook-derived outputs.
"""

from __future__ import annotations

import argparse
import base64
import concurrent.futures
import csv
import datetime as dt
import difflib
import hashlib
import json
import mimetypes
import os
import re
import ssl
import sys
import threading
import time
import traceback
import urllib.error
import urllib.request
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - exercised by user environment
    print(
        "Missing dependency: openpyxl. Install it with "
        "`python3 -m pip install openpyxl`.",
        file=sys.stderr,
    )
    raise SystemExit(2) from exc

try:
    import certifi
except ImportError:  # pragma: no cover - optional outside packaged app
    certifi = None


HEADER_ALIASES = {
    "test_id": ["TESTID", "Test ID", "TestID", "Test Id", "ID"],
    "category": ["Category"],
    "criterion": ["Criterion"],
    "weight": ["Weight"],
    "eval_method": ["Eval Method", "Type", "Test Type", "Prompt Type", "Output Type"],
    "applies_to": ["Applies To"],
    "prompt": ["Prompt", "Benchmark Prompt"],
    "input_material": [
        "Additional source information",
        "Additional Source Information",
        "Additional Source Info",
        "Source Information",
        "Input Material",
    ],
}
RUBRIC_HEADER_ALIASES = {
    "test_id": ["TESTID", "Test ID", "TestID", "Test Id", "ID"],
    "enabled": ["Enabled", "Run", "Include"],
}
PREFERRED_RUBRIC_SHEET_NAMES = (
    "Scoring Guide",
    "Rubric",
    "Rubric v3",
    "Model Testing Rubric",
)
REQUIRED_PROMPT_FIELDS = {"test_id", "prompt"}
REQUIRED_RUBRIC_FIELDS = {"test_id"}
OPTIONAL_PROMPT_DEFAULTS = {
    "category": "",
    "criterion": "",
    "weight": 1.0,
    "eval_method": "",
    "applies_to": "",
    "input_material": "",
}
MODEL_HEADER_ALIASES = {
    "key": ["Model Key", "Key", "ID", "Model ID Key", "model_id_string"],
    "name": ["Model Name", "Name", "Display Name", "name"],
    "model": [
        "OpenRouter Model ID",
        "API Model ID",
        "Provider Model ID",
        "Model ID",
        "Model",
        "Provider Model",
        "API Model",
        "model_id_string",
    ],
    "product": ["Product", "Tool", "Tool Name", "tool_name"],
    "provider": ["Provider", "Provider Key"],
    "provider_type": ["Provider Type", "Type"],
    "base_url": ["Base URL", "Endpoint", "API Base URL"],
    "api_key_env": ["API Key Env", "API Key Environment Variable", "Key Env"],
    "capabilities": ["Capabilities", "Capability", "Mode"],
    "enabled": ["Enabled", "Run", "Include"],
}
REQUIRED_MODEL_FIELDS = {"model"}
MANUAL_PROMPT_MARKERS = ("reviewer assessment",)
IMAGE_EDIT_PROMPT_MARKERS = (
    "replace ",
    "edit ",
    "modify ",
    "change ",
    "remove ",
    "retouch ",
    "restore ",
    "recolor ",
)
DEFAULT_MAX_TOKENS = 1000
CSV_COLUMNS = [
    "run_id",
    "timestamp",
    "model_key",
    "model_name",
    "provider",
    "provider_model",
    "test_id",
    "category",
    "criterion",
    "weight",
    "eval_method",
    "prompt_source",
    "input_source",
    "output_type",
    "response",
    "output_files",
    "output_urls",
    "latency_seconds",
    "prompt_tokens",
    "completion_tokens",
    "reasoning_tokens",
    "total_tokens",
    "usage",
    "error",
    "prompt_fingerprint",
    "rubric_fingerprint",
    "benchmark_fingerprint",
    "cache_status",
    "cache_source",
]


@dataclass(frozen=True)
class PromptTest:
    row_number: int
    test_id: str
    category: str
    criterion: str
    weight: float
    eval_method: str
    applies_to: str
    prompt: str
    input_material: str
    prompt_source: str = "cell"
    input_source: str = "cell"


@dataclass
class SkippedTest:
    test_id: str
    category: str
    criterion: str
    eval_method: str
    reason: str


class CreditLimitError(RuntimeError):
    """Provider says the account/key cannot afford the requested call."""


class RateLimitError(RuntimeError):
    """Provider rejected the request because a rate or usage limit was hit."""


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalized(value: str) -> str:
    return value.strip().lower().replace("—", "-").replace("–", "-")


def normalized_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.strip().lower())


def stable_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def fingerprint_value(value: Any) -> str:
    return hashlib.sha256(stable_json(value).encode("utf-8")).hexdigest()


def image_id(test_id: str) -> bool:
    value = test_id.strip().upper()
    return value.startswith(("IG", "IMG", "IMAGE"))


def evidence_id(test_id: str) -> bool:
    value = test_id.strip().upper()
    return value.startswith(("P&", "E"))


def is_image_test(test: PromptTest) -> bool:
    category = normalized(test.category)
    eval_method = normalized(test.eval_method)
    return (
        category == "image generation"
        or "image" in eval_method
        or image_id(test.test_id)
    )


def is_image_edit_test(test: PromptTest) -> bool:
    if not is_image_test(test):
        return False
    criterion = normalized(test.criterion)
    eval_method = normalized(test.eval_method)
    prompt = normalized(test.prompt)
    return (
        criterion == "image editing"
        or "image edit" in eval_method
        or prompt.startswith(IMAGE_EDIT_PROMPT_MARKERS)
    )


def parse_weight(value: Any) -> float:
    if value is None or value == "":
        return 1.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 1.0


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def default_config() -> dict[str, Any]:
    return {
        "providers": {
            "openrouter": {
                "type": "openai_compatible",
                "base_url": "https://openrouter.ai/api/v1",
                "api_key_env": "OPENROUTER_API_KEY",
                "extra_headers": {"X-Title": "AI Finder Model Evaluation"},
            },
            "openai": {
                "type": "openai_compatible",
                "base_url": "https://api.openai.com/v1",
                "api_key_env": "OPENAI_API_KEY",
                "max_tokens_param": "max_completion_tokens",
                "supports_custom_temperature": False,
            },
            "github_models": {
                "type": "openai_compatible",
                "base_url": "https://models.github.ai",
                "api_key_env": "GITHUB_MODELS_TOKEN",
                "chat_completions_path": "/inference/chat/completions",
                "extra_headers": {
                    "Accept": "application/vnd.github+json",
                    "X-GitHub-Api-Version": "2026-03-10",
                },
            },
            "openai_images": {
                "type": "openai_image_generation",
                "base_url": "https://api.openai.com/v1",
                "api_key_env": "OPENAI_API_KEY",
                "request": {
                    "image_count": 1,
                    "size": "1024x1024",
                    "timeout_seconds": 180,
                },
            },
        },
        "request": {
            "temperature": 0.2,
            "max_tokens": DEFAULT_MAX_TOKENS,
            "timeout_seconds": 120,
            "retries": 2,
            "sleep_seconds": 0.5,
        },
        "models": [],
    }


def truthy(value: Any, default: bool = True) -> bool:
    text = clean_text(value).lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "y", "run", "include", "enabled"}


def parse_capabilities(value: Any, provider: str, provider_type: str) -> list[str]:
    text = clean_text(value)
    if text:
        parts = re.split(r"[,;/|]+", text)
        capabilities = [normalized(part) for part in parts if clean_text(part)]
        if "both" in capabilities:
            return ["text", "image"]
        return capabilities
    if provider_type == "openai_image_generation" or "image" in normalized(provider):
        return ["image"]
    return ["text"]


def looks_like_audio_model(product: str, name: str, model_id: str) -> bool:
    haystack = normalized(" ".join([product, name, model_id]))
    return (
        "whisper" in haystack
        or model_id in {"large-v3", "large-v3-turbo", "small", "tiny"}
    )


def looks_like_openai_text_model(product: str, name: str, model_id: str) -> bool:
    haystack = normalized(" ".join([product, name, model_id]))
    return (
        "chatgpt" in haystack
        or "openai" in haystack
        or model_id.startswith(("gpt-", "o1", "o3", "o4"))
        or model_id.startswith("openai/")
    )


def openai_api_model_id(model_id: str) -> str:
    return model_id.removeprefix("openai/")


def unique_key(base: str, used: set[str]) -> str:
    root = safe_filename(base).lower() or "model"
    key = root
    index = 2
    while key in used:
        key = f"{root}-{index}"
        index += 1
    used.add(key)
    return key


def model_cell(ws: Any, row_number: int, columns: dict[str, int], field: str) -> Any:
    column = columns.get(field)
    return ws.cell(row_number, column).value if column else ""


def header_column_positions(ws: Any, header_row: int, aliases: list[str]) -> list[int]:
    values = [clean_text(cell.value) for cell in ws[header_row]]
    normalized_values = {
        normalized_header(value): index + 1
        for index, value in enumerate(values)
        if value
    }

    positions: list[int] = []
    for alias in aliases:
        column = normalized_values.get(normalized_header(alias))
        if column and column not in positions:
            positions.append(column)
    return positions


def first_non_empty_cell(ws: Any, row_number: int, columns: Iterable[int]) -> str:
    for column in columns:
        value = clean_text(ws.cell(row_number, column).value)
        if value:
            return value
    return ""


def workbook_defaults_to_openai(
    ws: Any,
    header_row: int,
    columns: dict[str, int],
    direct_model_columns: list[int],
    model_columns: list[int],
) -> bool:
    has_candidate = False
    for row_number in range(header_row + 1, ws.max_row + 1):
        provider = clean_text(model_cell(ws, row_number, columns, "provider"))
        provider_type = clean_text(model_cell(ws, row_number, columns, "provider_type"))
        capabilities_cell = model_cell(ws, row_number, columns, "capabilities")
        if provider or provider_type or clean_text(capabilities_cell):
            continue

        direct_model_id = first_non_empty_cell(ws, row_number, direct_model_columns)
        any_model_id = first_non_empty_cell(ws, row_number, model_columns)
        model_id = direct_model_id or any_model_id
        if not model_id:
            continue

        has_candidate = True
        name = clean_text(model_cell(ws, row_number, columns, "name"))
        product = clean_text(model_cell(ws, row_number, columns, "product"))
        if not looks_like_openai_text_model(product, name, model_id):
            return False

    return has_candidate


def read_model_workbook(models_path: Path, base_config: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(models_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, columns = find_header_row(ws, MODEL_HEADER_ALIASES, REQUIRED_MODEL_FIELDS)
    config = dict(base_config)
    config["providers"] = dict(base_config.get("providers", {}))
    config["models"] = []

    model_columns = header_column_positions(ws, header_row, MODEL_HEADER_ALIASES["model"])
    direct_model_columns = header_column_positions(ws, header_row, ["model_id_string"])
    default_text_provider = (
        "openai"
        if workbook_defaults_to_openai(
            ws, header_row, columns, direct_model_columns, model_columns
        )
        else "openrouter"
    )

    used_keys: set[str] = set()
    for row_number in range(header_row + 1, ws.max_row + 1):
        provider = clean_text(model_cell(ws, row_number, columns, "provider"))
        provider_type = clean_text(model_cell(ws, row_number, columns, "provider_type"))
        capabilities_cell = model_cell(ws, row_number, columns, "capabilities")
        raw_model_id = clean_text(model_cell(ws, row_number, columns, "model"))
        direct_model_id = first_non_empty_cell(ws, row_number, direct_model_columns)
        model_id = raw_model_id
        if not model_id and (provider or provider_type or clean_text(capabilities_cell)):
            model_id = first_non_empty_cell(ws, row_number, model_columns)
        if not model_id:
            continue

        name = clean_text(model_cell(ws, row_number, columns, "name")) or model_id
        product = clean_text(model_cell(ws, row_number, columns, "product"))
        if not clean_text(capabilities_cell) and looks_like_audio_model(product, name, model_id):
            capabilities = ["audio"]
        else:
            capabilities = parse_capabilities(capabilities_cell, provider, provider_type)
        if not provider:
            provider = "openai_images" if "image" in capabilities else default_text_provider
        if provider == "openai" and direct_model_id:
            model_id = direct_model_id
        if provider == "openai":
            model_id = openai_api_model_id(model_id)

        base_url = clean_text(model_cell(ws, row_number, columns, "base_url"))
        api_key_env = clean_text(model_cell(ws, row_number, columns, "api_key_env"))
        if base_url or api_key_env or provider_type:
            provider_config = dict(config["providers"].get(provider, {}))
            if provider_type:
                provider_config["type"] = provider_type
            elif "type" not in provider_config:
                provider_config["type"] = (
                    "openai_image_generation"
                    if "image" in capabilities
                    else "openai_compatible"
                )
            if base_url:
                provider_config["base_url"] = base_url
            if api_key_env:
                provider_config["api_key_env"] = api_key_env
            config["providers"][provider] = provider_config

        explicit_key = clean_text(model_cell(ws, row_number, columns, "key"))
        key = unique_key(explicit_key or name or model_id, used_keys)
        enabled = truthy(model_cell(ws, row_number, columns, "enabled"), default=True)
        config["models"].append(
            {
                "key": key,
                "name": name,
                "provider": provider,
                "model": model_id,
                "capabilities": capabilities,
                "enabled": enabled,
            }
        )

    if not config["models"]:
        raise ValueError("The models spreadsheet did not contain any model rows.")
    return config


def find_header_row(
    ws: Any,
    aliases: dict[str, list[str]],
    required_fields: set[str],
    max_scan_rows: int = 20,
) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [clean_text(cell.value) for cell in ws[row_number]]
        normalized_values = {
            normalized_header(value): index + 1
            for index, value in enumerate(values)
            if value
        }

        positions: dict[str, int] = {}
        for field, field_aliases in aliases.items():
            for alias in field_aliases:
                normalized_alias = normalized_header(alias)
                if normalized_alias in normalized_values:
                    positions[field] = normalized_values[normalized_alias]
                    break

        if required_fields.issubset(positions.keys()):
            return row_number, positions

    required = ", ".join(sorted(required_fields))
    raise ValueError(f"Could not find a header row containing: {required}")


def cell_value(ws: Any, row_number: int, columns: dict[str, int], field: str) -> Any:
    column = columns.get(field)
    if not column:
        return OPTIONAL_PROMPT_DEFAULTS.get(field, "")
    return ws.cell(row_number, column).value


def read_prompt_library(
    workbook_path: Path,
    sheet_name: str,
    inherit_shorthand: bool,
) -> list[PromptTest]:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        if sheet_name == "Test Prompts":
            sheet_name = wb.sheetnames[0]
        else:
            raise ValueError(
                f"Sheet {sheet_name!r} not found. Available sheets: {', '.join(wb.sheetnames)}"
            )

    ws = wb[sheet_name]
    header_row, columns = find_header_row(ws, HEADER_ALIASES, REQUIRED_PROMPT_FIELDS)
    tests: list[PromptTest] = []
    last_prompt = ""
    last_input = ""

    for row_number in range(header_row + 1, ws.max_row + 1):
        test_id = clean_text(ws.cell(row_number, columns["test_id"]).value)
        if not test_id:
            continue

        prompt = clean_text(cell_value(ws, row_number, columns, "prompt"))
        input_material = clean_text(cell_value(ws, row_number, columns, "input_material"))
        prompt_source = "cell"
        input_source = "cell"

        if inherit_shorthand and not prompt and last_prompt:
            prompt = last_prompt
            prompt_source = "inherited"
        if inherit_shorthand and input_material == "^" and last_input:
            input_material = last_input
            input_source = "inherited"

        if clean_text(ws.cell(row_number, columns["prompt"]).value):
            last_prompt = prompt
        if input_material and input_source == "cell" and input_material != "^":
            last_input = input_material

        category = clean_text(cell_value(ws, row_number, columns, "category"))
        criterion = clean_text(cell_value(ws, row_number, columns, "criterion"))
        eval_method = clean_text(cell_value(ws, row_number, columns, "eval_method"))
        if not eval_method and image_id(test_id):
            eval_method = "Prompt - Image"
        elif not eval_method:
            eval_method = "Prompt - Text"
        if not category and image_id(test_id):
            category = "Image Generation"
        elif not category:
            category = "Text"

        tests.append(
            PromptTest(
                row_number=row_number,
                test_id=test_id,
                category=category,
                criterion=criterion,
                weight=parse_weight(cell_value(ws, row_number, columns, "weight")),
                eval_method=eval_method,
                applies_to=clean_text(cell_value(ws, row_number, columns, "applies_to")),
                prompt=prompt,
                input_material=input_material,
                prompt_source=prompt_source,
                input_source=input_source,
            )
        )

    return tests


def read_rubric_test_fingerprints(workbook_path: Path, sheet_name: str | None) -> dict[str, str]:
    wb = load_workbook(workbook_path, data_only=True)
    candidate_sheets: list[Any] = []
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Rubric sheet {sheet_name!r} not found. Available sheets: {', '.join(wb.sheetnames)}"
            )
        candidate_sheets.append(wb[sheet_name])
    else:
        for preferred in PREFERRED_RUBRIC_SHEET_NAMES:
            if preferred in wb.sheetnames:
                candidate_sheets.append(wb[preferred])
        candidate_sheets.extend(
            wb[name] for name in wb.sheetnames if wb[name] not in candidate_sheets
        )

    for ws in candidate_sheets:
        try:
            header_row, columns = find_header_row(
                ws,
                RUBRIC_HEADER_ALIASES,
                REQUIRED_RUBRIC_FIELDS,
            )
        except ValueError:
            continue

        rows_by_test_id: dict[str, list[dict[str, str]]] = {}
        enabled_column = columns.get("enabled")
        header_values = {
            column: clean_text(ws.cell(header_row, column).value) or f"Column {column}"
            for column in range(1, ws.max_column + 1)
        }
        for row_number in range(header_row + 1, ws.max_row + 1):
            test_id = clean_text(ws.cell(row_number, columns["test_id"]).value)
            if not test_id:
                continue
            if enabled_column and not truthy(ws.cell(row_number, enabled_column).value, default=True):
                continue
            row_values = {
                header_values[column]: clean_text(ws.cell(row_number, column).value)
                for column in range(1, ws.max_column + 1)
                if clean_text(ws.cell(row_number, column).value)
            }
            rows_by_test_id.setdefault(test_id, []).append(
                {
                    "sheet": ws.title,
                    "row": str(row_number),
                    **row_values,
                }
            )

        if rows_by_test_id:
            return {
                test_id: fingerprint_value({"test_id": test_id, "rubric_rows": rows})
                for test_id, rows in rows_by_test_id.items()
            }

    raise ValueError(
        "Rubric workbook needs a sheet with a Test ID column. "
        f"Checked sheets: {', '.join(wb.sheetnames)}"
    )


def read_rubric_test_ids(workbook_path: Path, sheet_name: str | None) -> set[str]:
    return set(read_rubric_test_fingerprints(workbook_path, sheet_name))


def validate_rubric_coverage(
    tests: list[PromptTest],
    rubric_test_ids: set[str],
    allow_missing_rubric: bool,
) -> list[str]:
    missing = [
        test.test_id
        for test in tests
        if test.test_id and test.test_id not in rubric_test_ids
    ]
    if missing and not allow_missing_rubric:
        preview = ", ".join(missing[:25])
        suffix = f" ... and {len(missing) - 25} more" if len(missing) > 25 else ""
        raise ValueError(
            "Rubric workbook is missing enabled selected Test ID(s): "
            f"{preview}{suffix}. Update the rubric or pass --allow-missing-rubric."
        )
    return missing


def prompt_fingerprint(test: PromptTest) -> str:
    return fingerprint_value(
        {
            "schema": 1,
            "test_id": test.test_id,
            "category": test.category,
            "criterion": test.criterion,
            "weight": test.weight,
            "eval_method": test.eval_method,
            "applies_to": test.applies_to,
            "prompt": test.prompt,
            "input_material": test.input_material,
        }
    )


def model_fingerprint(model: dict[str, Any]) -> str:
    return fingerprint_value(
        {
            "schema": 1,
            "key": clean_text(model.get("key")),
            "provider": clean_text(model.get("provider")),
            "model": clean_text(model.get("model")),
        }
    )


def benchmark_fingerprint(
    model: dict[str, Any],
    test: PromptTest,
    rubric_fingerprints: dict[str, str],
) -> str:
    return fingerprint_value(
        {
            "schema": 1,
            "model": model_fingerprint(model),
            "prompt": prompt_fingerprint(test),
            "rubric": rubric_fingerprints.get(test.test_id, ""),
            "output_type": "image" if is_image_test(test) else "text",
        }
    )


def cacheable_result_row(row: dict[str, Any]) -> bool:
    if clean_text(row.get("error")):
        return False
    return bool(
        clean_text(row.get("response"))
        or clean_text(row.get("output_files"))
        or clean_text(row.get("output_urls"))
    )


def result_pair_key(row: dict[str, Any]) -> tuple[str, str]:
    return (
        clean_text(row.get("model_key")),
        clean_text(row.get("test_id")),
    )


def completed_result_row(row: dict[str, Any]) -> bool:
    return bool(
        clean_text(row.get("response"))
        or clean_text(row.get("output_files"))
        or clean_text(row.get("output_urls"))
        or clean_text(row.get("error"))
    )


def matching_existing_pair_keys(
    rows: Iterable[dict[str, Any]],
    pair_fingerprints: dict[tuple[str, str], str],
) -> set[tuple[str, str]]:
    keys: set[tuple[str, str]] = set()
    for row in rows:
        pair_key = result_pair_key(row)
        if pair_key not in pair_fingerprints or not completed_result_row(row):
            continue
        if clean_text(row.get("benchmark_fingerprint")) == pair_fingerprints[pair_key]:
            keys.add(pair_key)
    return keys


def cache_search_paths(history_dir: Path, output_dir: Path | None) -> list[Path]:
    if not history_dir.exists():
        return []

    output_dir_resolved = output_dir.resolve() if output_dir and output_dir.exists() else None
    paths: list[Path] = []
    for path in history_dir.rglob("responses.jsonl"):
        if output_dir_resolved:
            try:
                if path.resolve().is_relative_to(output_dir_resolved):
                    continue
            except OSError:
                pass
        paths.append(path)
    return sorted(paths)


def legacy_prompt_fingerprints(result_workbook_path: Path) -> dict[str, str]:
    if not result_workbook_path.exists():
        return {}
    try:
        tests = read_prompt_library(
            workbook_path=result_workbook_path,
            sheet_name="Test Prompts",
            inherit_shorthand=True,
        )
    except Exception:
        return {}
    return {test.test_id: prompt_fingerprint(test) for test in tests}


def history_prompt_fingerprints(
    history_dir: Path | None,
    output_dir: Path | None,
) -> dict[str, set[str]]:
    if not history_dir or not history_dir.exists():
        return {}

    output_dir_resolved = output_dir.resolve() if output_dir and output_dir.exists() else None
    fingerprints: dict[str, set[str]] = {}
    for path in sorted(history_dir.rglob("model_test_results.xlsx")):
        if output_dir_resolved:
            try:
                if path.resolve().is_relative_to(output_dir_resolved):
                    continue
            except OSError:
                pass
        for test_id, fingerprint in legacy_prompt_fingerprints(path).items():
            fingerprints.setdefault(test_id, set()).add(fingerprint)
    return fingerprints


def changed_test_ids(
    tests: Iterable[PromptTest],
    prompt_fingerprints_by_test_id: dict[str, str],
    historical_prompt_fingerprints: dict[str, set[str]],
) -> set[str]:
    changed: set[str] = set()
    for test in tests:
        current = prompt_fingerprints_by_test_id.get(test.test_id)
        if not current:
            continue
        if current not in historical_prompt_fingerprints.get(test.test_id, set()):
            changed.add(test.test_id)
    return changed


def read_cached_result_rows(
    history_dir: Path | None,
    output_dir: Path | None,
    needed_fingerprints: set[str],
    pair_fingerprints: dict[tuple[str, str], str] | None = None,
    prompt_fingerprints_by_test_id: dict[str, str] | None = None,
    rubric_fingerprints: dict[str, str] | None = None,
    model_identity_by_key: dict[str, tuple[str, str]] | None = None,
) -> dict[str, dict[str, Any]]:
    if not history_dir or not needed_fingerprints:
        return {}

    cached: dict[str, dict[str, Any]] = {}
    cached_timestamps: dict[str, str] = {}
    legacy_prompt_cache: dict[Path, dict[str, str]] = {}
    for path in cache_search_paths(history_dir, output_dir):
        try:
            rows = load_existing_jsonl(path)
        except (OSError, json.JSONDecodeError):
            continue
        for row in rows:
            fingerprint = clean_text(row.get("benchmark_fingerprint"))
            pair_key = (
                clean_text(row.get("model_key")),
                clean_text(row.get("test_id")),
            )
            if not fingerprint and pair_fingerprints and pair_key in pair_fingerprints:
                result_workbook_path = path.parent / "model_test_results.xlsx"
                if result_workbook_path not in legacy_prompt_cache:
                    legacy_prompt_cache[result_workbook_path] = legacy_prompt_fingerprints(
                        result_workbook_path
                    )
                legacy_prompt = legacy_prompt_cache[result_workbook_path].get(pair_key[1])
                current_prompt = (prompt_fingerprints_by_test_id or {}).get(pair_key[1])
                current_model = (model_identity_by_key or {}).get(pair_key[0], ("", ""))
                if (
                    legacy_prompt
                    and current_prompt
                    and legacy_prompt == current_prompt
                    and (
                        not current_model[0]
                        or clean_text(row.get("provider_model")) == current_model[0]
                    )
                    and (
                        not current_model[1]
                        or clean_text(row.get("provider")) == current_model[1]
                    )
                ):
                    fingerprint = pair_fingerprints[pair_key]
                    row = {
                        **row,
                        "prompt_fingerprint": current_prompt,
                        "rubric_fingerprint": (rubric_fingerprints or {}).get(pair_key[1], ""),
                        "benchmark_fingerprint": fingerprint,
                    }
            if fingerprint not in needed_fingerprints or not cacheable_result_row(row):
                continue
            timestamp = clean_text(row.get("timestamp"))
            if fingerprint in cached and timestamp <= cached_timestamps.get(fingerprint, ""):
                continue
            cached_row = dict(row)
            cached_row["_cache_source_path"] = str(path)
            cached[fingerprint] = cached_row
            cached_timestamps[fingerprint] = timestamp
    return cached


def reusable_result_row(
    cached_row: dict[str, Any],
    run_id: str,
    cache_source: str,
) -> dict[str, Any]:
    row = {column: cached_row.get(column, "") for column in CSV_COLUMNS}
    row["run_id"] = run_id
    row["timestamp"] = dt.datetime.now(dt.timezone.utc).isoformat()
    row["latency_seconds"] = 0
    row["prompt_tokens"] = ""
    row["completion_tokens"] = ""
    row["reasoning_tokens"] = ""
    row["total_tokens"] = ""
    row["usage"] = {}
    row["cache_status"] = "reused"
    row["cache_source"] = cache_source
    return row


def skip_reason(test: PromptTest, args: argparse.Namespace) -> str | None:
    category = normalized(test.category)
    eval_method = normalized(test.eval_method)
    prompt = normalized(test.prompt)

    if not test.prompt:
        return "missing prompt"
    if not args.include_image and is_image_test(test):
        return "image prompt skipped"
    if args.include_image and is_image_edit_test(test) and not test.input_material:
        return "image edit skipped; no source image path in input material"
    if not args.include_evidence and (
        eval_method == "evidence review"
        or category in {"privacy & data safety", "enterprise"}
        or evidence_id(test.test_id)
    ):
        return "evidence/security review skipped"
    if not args.include_manual_review and any(
        marker in prompt for marker in MANUAL_PROMPT_MARKERS
    ):
        return "manual reviewer assessment skipped"
    if not args.include_manual_review and normalized(test.applies_to).startswith(
        "per product"
    ):
        return "product-level manual review skipped"
    return None


def split_filter(value: str | None) -> set[str]:
    if not value:
        return set()
    return {part.strip() for part in value.split(",") if part.strip()}


def website_score_fields(headers: Iterable[str]) -> list[str]:
    return [
        header
        for header in headers
        if header.startswith("score_") or header.endswith("_score")
    ]


def read_scored_model_keys(seed_csv_path: Path) -> dict[str, list[str]]:
    with seed_csv_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        headers = reader.fieldnames or []
        score_fields = website_score_fields(headers)
        scored: dict[str, list[str]] = {}
        for row in reader:
            model_key = clean_text(row.get("model_id_string"))
            if not model_key:
                continue
            populated = [
                field
                for field in score_fields
                if clean_text(row.get(field))
            ]
            if populated:
                scored[model_key] = populated
        return scored


def filter_scored_models(
    models: list[dict[str, Any]],
    scored_model_keys: dict[str, list[str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    selected: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    for model in models:
        model_key = clean_text(model.get("key"))
        if model_key and model_key in scored_model_keys:
            skipped.append(model)
        else:
            selected.append(model)
    return selected, skipped


def eligible_tests(
    tests: Iterable[PromptTest],
    args: argparse.Namespace,
) -> tuple[list[PromptTest], list[SkippedTest]]:
    only_tests = split_filter(args.only_tests)
    selected: list[PromptTest] = []
    skipped: list[SkippedTest] = []

    for test in tests:
        if only_tests and test.test_id not in only_tests:
            continue

        reason = skip_reason(test, args)
        if reason:
            skipped.append(
                SkippedTest(
                    test_id=test.test_id,
                    category=test.category,
                    criterion=test.criterion,
                    eval_method=test.eval_method,
                    reason=reason,
                )
            )
            continue

        selected.append(test)

    if args.limit is not None:
        limited = selected[: args.limit]
        skipped.extend(
            SkippedTest(
                test_id=test.test_id,
                category=test.category,
                criterion=test.criterion,
                eval_method=test.eval_method,
                reason="outside --limit",
            )
            for test in selected[args.limit :]
        )
        selected = limited

    return selected, skipped


def enabled_models(config: dict[str, Any], only_models: str | None) -> list[dict[str, Any]]:
    selected_keys = split_filter(only_models)
    models = [
        model
        for model in config.get("models", [])
        if model.get("enabled", True) and (not selected_keys or model.get("key") in selected_keys)
    ]
    if selected_keys:
        missing = selected_keys - {model.get("key") for model in models}
        if missing:
            raise ValueError(f"Requested model key(s) not enabled/found: {', '.join(sorted(missing))}")
    return models


def provider_for(config: dict[str, Any], model: dict[str, Any]) -> dict[str, Any]:
    providers = config.get("providers", {})
    provider_key = model.get("provider")
    provider = providers.get(provider_key)
    if not provider:
        raise ValueError(f"Unknown provider {provider_key!r} for model {model.get('key')!r}.")
    return provider


def model_capabilities(config: dict[str, Any], model: dict[str, Any]) -> set[str]:
    configured = model.get("capabilities")
    if configured:
        if isinstance(configured, str):
            configured = [configured]
        return {normalized(str(capability)) for capability in configured}

    provider = provider_for(config, model)
    provider_type = provider.get("type", "openai_compatible")
    if provider_type == "openai_image_generation":
        return {"image"}
    return {"text"}


def model_supports_test(config: dict[str, Any], model: dict[str, Any], test: PromptTest) -> bool:
    required = "image" if is_image_test(test) else "text"
    capabilities = model_capabilities(config, model)
    return required in capabilities or "both" in capabilities


def planned_pairs(
    config: dict[str, Any],
    models: list[dict[str, Any]],
    tests: list[PromptTest],
) -> list[tuple[dict[str, Any], PromptTest]]:
    return [
        (model, test)
        for model in models
        for test in tests
        if model_supports_test(config, model, test)
    ]


def product_lane_key(config: dict[str, Any], model: dict[str, Any]) -> str:
    provider_key = clean_text(model.get("provider")) or "unknown"
    provider = provider_for(config, model)
    base_url = str(provider.get("base_url", "")).lower()
    model_id = clean_text(model.get("model"))

    if "api.openai.com" in base_url or provider_key in {"openai", "openai_images"}:
        return "openai"
    if openrouter_provider(provider) and "/" in model_id:
        return model_id.split("/", 1)[0]
    return provider_key


def product_lane_label(key: str) -> str:
    labels = {
        "anthropic": "Anthropic",
        "deepseek": "DeepSeek",
        "google": "Google",
        "meta-llama": "Meta Llama",
        "mistralai": "Mistral",
        "openai": "OpenAI",
    }
    return labels.get(key, key)


def group_pairs_by_product(
    config: dict[str, Any],
    pairs: list[tuple[dict[str, Any], PromptTest]],
) -> dict[str, list[tuple[dict[str, Any], PromptTest]]]:
    groups: dict[str, list[tuple[dict[str, Any], PromptTest]]] = {}
    for model, test in pairs:
        groups.setdefault(product_lane_key(config, model), []).append((model, test))
    return groups


def product_worker_count(
    parallel_products: bool,
    requested_workers: int,
    lane_count: int,
) -> int:
    if not parallel_products:
        return 1
    if lane_count <= 1:
        return 1
    if requested_workers:
        return max(1, min(requested_workers, lane_count))
    return lane_count


def unsupported_tests(
    config: dict[str, Any],
    models: list[dict[str, Any]],
    tests: list[PromptTest],
) -> list[PromptTest]:
    return [
        test
        for test in tests
        if not any(model_supports_test(config, model, test) for model in models)
    ]


def required_api_key_envs(
    config: dict[str, Any],
    pairs: list[tuple[dict[str, Any], PromptTest]],
) -> set[str]:
    envs: set[str] = set()
    for model, _test in pairs:
        provider = provider_for(config, model)
        env_name = clean_text(provider.get("api_key_env"))
        if env_name:
            envs.add(env_name)
    return envs


def validate_api_keys(
    config: dict[str, Any],
    pairs: list[tuple[dict[str, Any], PromptTest]],
) -> None:
    missing = sorted(
        env_name for env_name in required_api_key_envs(config, pairs) if not os.getenv(env_name)
    )
    if missing:
        raise ValueError(
            "Missing API key environment variable(s): "
            f"{', '.join(missing)}. "
            "Use Dry Run to validate without API calls, enter the key in the SwiftUI app, "
            "or export it before running the CLI."
        )


def certifi_https_context() -> ssl.SSLContext | None:
    if certifi is None:
        return None
    ca_bundle = Path(certifi.where())
    if not ca_bundle.exists():
        return None
    return ssl.create_default_context(cafile=str(ca_bundle))


def open_url(url_or_request: str | urllib.request.Request, timeout: int) -> Any:
    context = certifi_https_context()
    if context is None:
        return urllib.request.urlopen(url_or_request, timeout=timeout)
    return urllib.request.urlopen(url_or_request, timeout=timeout, context=context)


def openrouter_model_catalogue(base_url: str) -> set[str]:
    url = f"{base_url.rstrip('/')}/models"
    with open_url(url, timeout=30) as response:
        payload = json.loads(response.read().decode("utf-8"))
    return {
        clean_text(item.get("id"))
        for item in payload.get("data", [])
        if clean_text(item.get("id"))
    }


def openrouter_provider(provider: dict[str, Any]) -> bool:
    return "openrouter.ai" in str(provider.get("base_url", "")).lower()


def dot_version_variants(model_id: str) -> list[str]:
    variants = {model_id}
    variants.add(re.sub(r"-(\d+)-(\d+)$", r"-\1.\2", model_id))
    variants.add(re.sub(r"-(\d+)-(\d+)(-)", r"-\1.\2\3", model_id))
    variants.add(model_id.replace("-preview", "-preview"))
    return [variant for variant in variants if variant != model_id]


def likely_openrouter_ids(model_id: str, catalogue: set[str]) -> list[str]:
    candidates: list[str] = []
    possible = {model_id}
    possible.update(dot_version_variants(model_id))

    provider_prefixes = []
    if model_id.startswith("gpt-"):
        provider_prefixes.append("openai")
    if model_id.startswith("claude-"):
        provider_prefixes.append("anthropic")
    if model_id.startswith("gemini-"):
        provider_prefixes.append("google")
    if model_id.startswith("deepseek-"):
        provider_prefixes.append("deepseek")
    if model_id.startswith("llama-"):
        provider_prefixes.append("meta-llama")
    if model_id.startswith("mistral-"):
        provider_prefixes.append("mistralai")

    for value in list(possible):
        candidates.extend([item for item in catalogue if item.endswith(f"/{value}")])
        for prefix in provider_prefixes:
            candidates.append(f"{prefix}/{value}")

    exact = [candidate for candidate in candidates if candidate in catalogue]
    if exact:
        return sorted(set(exact))[:5]

    model_words = re.sub(r"[^a-z0-9]+", " ", model_id.lower()).split()
    if model_words:
        contains_words = [
            item
            for item in catalogue
            if all(word in item.lower() for word in model_words[:4])
        ]
        if contains_words:
            return sorted(contains_words)[:5]

    close = difflib.get_close_matches(model_id, sorted(catalogue), n=5, cutoff=0.62)
    return close


def validate_openrouter_model_ids(
    config: dict[str, Any],
    pairs: list[tuple[dict[str, Any], PromptTest]],
) -> None:
    provider_models: dict[str, tuple[dict[str, Any], set[str]]] = {}
    for model, _test in pairs:
        provider = provider_for(config, model)
        if not openrouter_provider(provider):
            continue
        base_url = str(provider.get("base_url", "")).rstrip("/")
        provider_models.setdefault(base_url, (provider, set()))[1].add(clean_text(model.get("model")))

    errors: list[str] = []
    for base_url, (_provider, model_ids) in provider_models.items():
        catalogue = openrouter_model_catalogue(base_url)
        invalid = sorted(model_id for model_id in model_ids if model_id not in catalogue)
        if not invalid:
            continue
        lines = [f"Invalid OpenRouter model ID(s) for {base_url}:"]
        for model_id in invalid:
            suggestions = likely_openrouter_ids(model_id, catalogue)
            hint = f" Try: {', '.join(suggestions)}" if suggestions else ""
            lines.append(f"  - {model_id}.{hint}")
        lines.append(
            "Add an 'OpenRouter Model ID' column with the exact slug from "
            "https://openrouter.ai/models, or update model_id_string to the exact slug."
        )
        errors.append("\n".join(lines))

    if errors:
        raise ValueError("\n\n".join(errors))


def create_messages(test: PromptTest) -> list[dict[str, str]]:
    user_parts = [test.prompt]
    if test.input_material:
        user_parts.append(f"Input material:\n{test.input_material}")
    return [
        {
            "role": "system",
            "content": (
                "You are responding to a benchmark prompt. Answer the user's "
                "request directly and do not mention the benchmark unless asked."
            ),
        },
        {"role": "user", "content": "\n\n".join(user_parts)},
    ]


def normalize_content(content: Any) -> str:
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts: list[str] = []
        for item in content:
            if isinstance(item, dict):
                if item.get("type") == "text" and item.get("text"):
                    parts.append(str(item["text"]))
                elif item.get("content"):
                    parts.append(str(item["content"]))
            elif item:
                parts.append(str(item))
        return "\n".join(parts)
    return "" if content is None else str(content)


def merge_request_options(
    config: dict[str, Any],
    provider: dict[str, Any],
    model: dict[str, Any],
) -> dict[str, Any]:
    options: dict[str, Any] = {}
    options.update(config.get("request", {}))
    options.update(provider.get("request", {}))
    options.update(model.get("request", {}))
    return options


def max_tokens_parameter(provider: dict[str, Any], model: dict[str, Any]) -> str:
    configured = clean_text(model.get("max_tokens_param") or provider.get("max_tokens_param"))
    if configured:
        return configured

    base_url = str(provider.get("base_url", "")).lower()
    if "api.openai.com" in base_url:
        return "max_completion_tokens"
    return "max_tokens"


def supports_custom_temperature(provider: dict[str, Any], model: dict[str, Any]) -> bool:
    configured = model.get(
        "supports_custom_temperature",
        provider.get("supports_custom_temperature"),
    )
    if configured is not None:
        return truthy(configured, default=True)

    base_url = str(provider.get("base_url", "")).lower()
    return "api.openai.com" not in base_url


def redact_secrets(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"Bearer\s+[^'\"\\\s]+", "Bearer [redacted]", text)
    text = re.sub(r"github_pat_[A-Za-z0-9_]+", "github_pat_[redacted]", text)
    text = re.sub(r"gh[opsru]_[A-Za-z0-9_]+", "gh[redacted]", text)
    text = re.sub(r"sk-or-v1-[A-Za-z0-9]+", "sk-or-v1-[redacted]", text)
    text = re.sub(r"sk-proj-[A-Za-z0-9_-]+", "sk-proj-[redacted]", text)
    text = re.sub(r"sk-[A-Za-z0-9][A-Za-z0-9_-]{12,}", "sk-[redacted]", text)
    return text


def api_key_for(provider: dict[str, Any]) -> str:
    env_name = provider.get("api_key_env")
    if not env_name:
        return ""
    api_key = re.sub(r"\s+", "", os.getenv(env_name) or "")
    if not api_key:
        raise ValueError(f"Environment variable {env_name} is not set.")
    return api_key


def header_value(value: Any) -> str:
    return re.sub(r"[\r\n]+", " ", clean_text(value)).strip()


def build_headers(provider: dict[str, Any], model: dict[str, Any]) -> dict[str, str]:
    headers = {
        "Content-Type": "application/json",
    }
    api_key = api_key_for(provider)
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    headers.update(
        {
            header_value(name): header_value(value)
            for name, value in provider.get("extra_headers", {}).items()
        }
    )
    headers.update(
        {
            header_value(name): header_value(value)
            for name, value in model.get("extra_headers", {}).items()
        }
    )
    return headers


def compact_http_error_body(body: str) -> str:
    try:
        parsed = json.loads(body)
    except json.JSONDecodeError:
        return body.strip()

    error = parsed.get("error") if isinstance(parsed, dict) else None
    if isinstance(error, dict):
        message = clean_text(error.get("message"))
        code = clean_text(error.get("code"))
        error_type = clean_text(error.get("type"))
        details = [part for part in [message, f"code={code}" if code else "", f"type={error_type}" if error_type else ""] if part]
        return "; ".join(details) or body.strip()
    return body.strip()


def retry_delay_seconds(exc: urllib.error.HTTPError, attempt: int) -> float:
    headers = exc.headers
    for name in ("retry-after-ms", "Retry-After-Ms"):
        value = clean_text(headers.get(name))
        if value:
            try:
                return max(0.0, float(value) / 1000.0)
            except ValueError:
                pass

    for name in ("retry-after", "Retry-After"):
        value = clean_text(headers.get(name))
        if value:
            try:
                return max(0.0, float(value))
            except ValueError:
                pass

    return min(30.0, float(2**attempt))


def http_error(exc: urllib.error.HTTPError, body: str) -> RuntimeError:
    message = f"HTTP {exc.code}: {compact_http_error_body(body)}"
    if exc.code == 429:
        return RateLimitError(message)
    if exc.code == 402:
        return CreditLimitError(message)
    return RuntimeError(message)


def post_json(
    provider: dict[str, Any],
    model: dict[str, Any],
    path: str,
    payload: dict[str, Any],
    timeout: int,
    retries: int,
) -> dict[str, Any]:
    base_url = str(provider.get("base_url", "")).rstrip("/")
    if not base_url:
        raise ValueError("Provider is missing base_url.")

    data = json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        f"{base_url}{path}",
        data=data,
        headers=build_headers(provider, model),
    )
    last_error: Exception | None = None

    for attempt in range(retries + 1):
        retry_delay = min(8.0, float(2**attempt))
        try:
            with open_url(request, timeout=timeout) as response:
                body = response.read().decode("utf-8")
            return json.loads(body)
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            last_error = http_error(exc, body)
            if exc.code == 429:
                retry_delay = retry_delay_seconds(exc, attempt)
            if exc.code < 500 and exc.code != 429:
                raise last_error from exc
        except (urllib.error.URLError, TimeoutError) as exc:
            last_error = exc

        if attempt < retries:
            time.sleep(retry_delay)

    if isinstance(last_error, (CreditLimitError, RateLimitError)):
        raise last_error
    raise RuntimeError(f"API request failed after retries: {last_error}") from last_error


def call_openai_compatible(
    config: dict[str, Any],
    model: dict[str, Any],
    messages: list[dict[str, str]],
) -> tuple[str, dict[str, Any]]:
    provider = provider_for(config, model)

    provider_type = provider.get("type", "openai_compatible")
    if provider_type != "openai_compatible":
        raise ValueError(f"Unsupported provider type: {provider_type}")

    base_url = str(provider.get("base_url", "")).rstrip("/")
    if not base_url:
        raise ValueError(f"Provider {provider_key!r} is missing base_url.")

    options = merge_request_options(config, provider, model)
    max_tokens_key = max_tokens_parameter(provider, model)
    requested_max_tokens = options.get("max_tokens", DEFAULT_MAX_TOKENS)
    payload: dict[str, Any] = {
        "model": model.get("model"),
        "messages": messages,
        max_tokens_key: requested_max_tokens,
    }
    if supports_custom_temperature(provider, model):
        payload["temperature"] = options.get("temperature", 0.2)
    reasoning_effort = clean_text(options.get("reasoning_effort")).lower()
    if reasoning_effort and openrouter_provider(provider):
        payload["reasoning"] = {"effort": reasoning_effort}
    payload.update(provider.get("extra_body", {}))
    payload.update(model.get("extra_body", {}))

    path = provider.get("chat_completions_path", "/chat/completions")
    timeout = int(options.get("timeout_seconds", 120))
    retries = int(options.get("retries", 2))

    parsed = post_json(provider, model, path, payload, timeout, retries)
    choice = parsed.get("choices", [{}])[0]
    message = choice.get("message", {})
    usage = parsed.get("usage", {})
    if not isinstance(usage, dict):
        usage = {}
    else:
        usage = dict(usage)
    finish_reason = clean_text(choice.get("finish_reason"))
    if finish_reason:
        usage["_finish_reason"] = finish_reason
    usage["_requested_max_tokens"] = requested_max_tokens
    return normalize_content(message.get("content")), usage


def safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "-", value).strip("-") or "item"


def image_prompt(test: PromptTest) -> str:
    parts = [test.prompt]
    if test.input_material and not input_image_path(test):
        parts.append(f"Input material:\n{test.input_material}")
    return "\n\n".join(parts)


def input_image_path(test: PromptTest) -> Path | None:
    raw = clean_text(test.input_material)
    if not raw:
        return None

    candidate = Path(raw).expanduser()
    if not candidate.is_absolute():
        candidate = Path.cwd() / candidate
    return candidate if candidate.exists() and candidate.is_file() else None


def image_payload_options(options: dict[str, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    option_map = {
        "image_count": "n",
        "n": "n",
        "size": "size",
        "quality": "quality",
        "style": "style",
        "response_format": "response_format",
        "background": "background",
        "moderation": "moderation",
        "output_format": "output_format",
    }
    for source, target in option_map.items():
        if source in options and options[source] is not None:
            payload[target] = options[source]
    return payload


def image_file_extension(options: dict[str, Any]) -> str:
    extension = normalized(str(options.get("output_format") or "png"))
    if extension == "jpeg":
        return "jpg"
    if extension in {"jpg", "png", "webp"}:
        return extension
    return "png"


def save_image_outputs(
    parsed: dict[str, Any],
    output_dir: Path,
    model: dict[str, Any],
    test: PromptTest,
    file_extension: str = "png",
) -> tuple[list[str], list[str], str]:
    image_dir = output_dir / "images" / safe_filename(str(model.get("key", "model")))
    image_dir.mkdir(parents=True, exist_ok=True)

    files: list[str] = []
    urls: list[str] = []
    notes: list[str] = []
    data_items = parsed.get("data", [])

    for index, item in enumerate(data_items, start=1):
        if not isinstance(item, dict):
            continue
        revised_prompt = clean_text(item.get("revised_prompt"))
        if revised_prompt:
            notes.append(f"Revised prompt {index}: {revised_prompt}")

        if item.get("b64_json"):
            image_bytes = base64.b64decode(item["b64_json"])
            output_path = (
                image_dir
                / (
                    f"{safe_filename(test.test_id)}-"
                    f"{safe_filename(str(model.get('key', 'model')))}-"
                    f"{index}.{file_extension}"
                )
            )
            output_path.write_bytes(image_bytes)
            files.append(str(output_path.resolve()))
        if item.get("url"):
            urls.append(str(item["url"]))

    if not files and not urls:
        raise RuntimeError("Image response did not include b64_json or url data.")

    response = "Generated image output."
    if notes:
        response = f"{response}\n" + "\n".join(notes)
    return files, urls, response


def encode_multipart(
    fields: dict[str, Any],
    files: list[tuple[str, Path]],
) -> tuple[bytes, str]:
    boundary = f"----codex-model-eval-{uuid.uuid4().hex}"
    body = bytearray()

    def add(value: bytes) -> None:
        body.extend(value)

    for name, value in fields.items():
        add(f"--{boundary}\r\n".encode("utf-8"))
        add(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode("utf-8"))
        add(str(value).encode("utf-8"))
        add(b"\r\n")

    for name, path in files:
        mime_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        add(f"--{boundary}\r\n".encode("utf-8"))
        add(
            (
                f'Content-Disposition: form-data; name="{name}"; '
                f'filename="{path.name}"\r\n'
            ).encode("utf-8")
        )
        add(f"Content-Type: {mime_type}\r\n\r\n".encode("utf-8"))
        add(path.read_bytes())
        add(b"\r\n")

    add(f"--{boundary}--\r\n".encode("utf-8"))
    return bytes(body), f"multipart/form-data; boundary={boundary}"


def post_multipart(
    provider: dict[str, Any],
    model: dict[str, Any],
    path: str,
    fields: dict[str, Any],
    files: list[tuple[str, Path]],
    timeout: int,
    retries: int,
) -> dict[str, Any]:
    base_url = str(provider.get("base_url", "")).rstrip("/")
    if not base_url:
        raise ValueError("Provider is missing base_url.")

    body, content_type = encode_multipart(fields, files)
    headers = build_headers(provider, model)
    headers["Content-Type"] = content_type
    request = urllib.request.Request(f"{base_url}{path}", data=body, headers=headers)
    last_error: Exception | None = None

    for attempt in range(retries + 1):
        retry_delay = min(8.0, float(2**attempt))
        try:
            with open_url(request, timeout=timeout) as response:
                response_body = response.read().decode("utf-8")
            return json.loads(response_body)
        except urllib.error.HTTPError as exc:
            response_body = exc.read().decode("utf-8", errors="replace")
            last_error = http_error(exc, response_body)
            if exc.code == 429:
                retry_delay = retry_delay_seconds(exc, attempt)
            if exc.code < 500 and exc.code != 429:
                raise last_error from exc
        except (urllib.error.URLError, TimeoutError) as exc:
            last_error = exc

        if attempt < retries:
            time.sleep(retry_delay)

    if isinstance(last_error, (CreditLimitError, RateLimitError)):
        raise last_error
    raise RuntimeError(f"API request failed after retries: {last_error}") from last_error


def call_openai_image_generation(
    config: dict[str, Any],
    model: dict[str, Any],
    test: PromptTest,
    output_dir: Path,
) -> tuple[str, list[str], list[str], dict[str, Any]]:
    provider = provider_for(config, model)
    provider_type = provider.get("type", "openai_compatible")
    if provider_type != "openai_image_generation":
        raise ValueError(f"Unsupported image provider type: {provider_type}")

    options = merge_request_options(config, provider, model)
    timeout = int(options.get("timeout_seconds", 180))
    retries = int(options.get("retries", 2))
    payload = {
        "model": model.get("model"),
        "prompt": image_prompt(test),
        **image_payload_options(options),
    }
    payload.update(provider.get("extra_body", {}))
    payload.update(model.get("extra_body", {}))

    path = provider.get("image_generations_path", "/images/generations")
    parsed = post_json(provider, model, path, payload, timeout, retries)
    files, urls, response = save_image_outputs(
        parsed,
        output_dir,
        model,
        test,
        file_extension=image_file_extension(options),
    )
    return response, files, urls, parsed.get("usage", {})


def call_openai_image_edit(
    config: dict[str, Any],
    model: dict[str, Any],
    test: PromptTest,
    output_dir: Path,
) -> tuple[str, list[str], list[str], dict[str, Any]]:
    provider = provider_for(config, model)
    provider_type = provider.get("type", "openai_compatible")
    if provider_type != "openai_image_generation":
        raise ValueError(f"Unsupported image provider type: {provider_type}")

    source_image = input_image_path(test)
    if not source_image:
        raise ValueError("Image edit test requires input_material to be a local image path.")

    options = merge_request_options(config, provider, model)
    timeout = int(options.get("timeout_seconds", 180))
    retries = int(options.get("retries", 2))
    fields = {
        "model": model.get("model"),
        "prompt": test.prompt,
        **image_payload_options(options),
    }
    fields.update(provider.get("extra_body", {}))
    fields.update(model.get("extra_body", {}))

    path = provider.get("image_edits_path", "/images/edits")
    parsed = post_multipart(
        provider=provider,
        model=model,
        path=path,
        fields=fields,
        files=[("image", source_image)],
        timeout=timeout,
        retries=retries,
    )
    files, urls, response = save_image_outputs(
        parsed,
        output_dir,
        model,
        test,
        file_extension=image_file_extension(options),
    )
    return response, files, urls, parsed.get("usage", {})


def call_image_model(
    config: dict[str, Any],
    model: dict[str, Any],
    test: PromptTest,
    output_dir: Path,
) -> tuple[str, list[str], list[str], dict[str, Any]]:
    if is_image_edit_test(test):
        return call_openai_image_edit(config, model, test, output_dir)
    return call_openai_image_generation(config, model, test, output_dir)


def load_existing_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []

    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def append_jsonl(path: Path, row: dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def make_output_dir(base: Path | None) -> Path:
    if base:
        base.mkdir(parents=True, exist_ok=True)
        return base

    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    output_dir = Path("outputs") / "model_tests" / stamp
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def stringify_usage(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, sort_keys=True)


def token_count_value(value: Any) -> int | str:
    if value in (None, "") or isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return ""


def usage_token_count(usage: dict[str, Any], key: str) -> int | str:
    return token_count_value(usage.get(key))


def usage_detail_token_count(
    usage: dict[str, Any],
    details_key: str,
    key: str,
) -> int | str:
    details = usage.get(details_key)
    if not isinstance(details, dict):
        return ""
    return token_count_value(details.get(key))


def reasoning_token_count(usage: dict[str, Any]) -> int | str:
    top_level = usage_token_count(usage, "reasoning_tokens")
    if top_level != "":
        return top_level
    return usage_detail_token_count(
        usage,
        "completion_tokens_details",
        "reasoning_tokens",
    )


def row_int_value(row: dict[str, Any], key: str) -> int | None:
    value = token_count_value(row.get(key))
    return value if isinstance(value, int) else None


def row_usage(row: dict[str, Any]) -> dict[str, Any]:
    usage = row.get("usage")
    if isinstance(usage, dict):
        return usage
    if isinstance(usage, str) and usage.strip():
        try:
            parsed = json.loads(usage)
        except json.JSONDecodeError:
            return {}
        return parsed if isinstance(parsed, dict) else {}
    return {}


def finish_reason_hit_token_limit(finish_reason: str) -> bool:
    normalized_reason = normalized(finish_reason).replace("_", " ")
    return normalized_reason in {
        "length",
        "max tokens",
        "max completion tokens",
        "token limit",
        "tokens",
    }


def token_budget_warning(row: dict[str, Any]) -> str:
    if clean_text(row.get("output_type")) == "image" or clean_text(row.get("error")):
        return ""

    usage = row_usage(row)
    requested_max_tokens_value = token_count_value(usage.get("_requested_max_tokens"))
    requested_max_tokens = (
        requested_max_tokens_value if isinstance(requested_max_tokens_value, int) else None
    )
    completion_tokens = row_int_value(row, "completion_tokens")
    reasoning_tokens = row_int_value(row, "reasoning_tokens")
    finish_reason = clean_text(usage.get("_finish_reason"))
    hit_finish_limit = finish_reason_hit_token_limit(finish_reason)
    hit_token_cap = (
        requested_max_tokens is not None
        and completion_tokens is not None
        and completion_tokens >= requested_max_tokens
    )
    if not hit_finish_limit and not hit_token_cap:
        return ""

    has_visible_output = bool(
        clean_text(row.get("response"))
        or clean_text(row.get("output_files"))
        or clean_text(row.get("output_urls"))
    )
    details = []
    if completion_tokens is not None:
        details.append(f"completion_tokens={completion_tokens}")
    if reasoning_tokens is not None:
        details.append(f"reasoning_tokens={reasoning_tokens}")
    if requested_max_tokens is not None:
        details.append(f"max_tokens={requested_max_tokens}")
    if finish_reason:
        details.append(f"finish_reason={finish_reason}")
    detail_text = "; ".join(details)
    if has_visible_output:
        return f"response likely truncated after hitting the token cap ({detail_text})"
    return f"blank response likely exhausted the token cap before visible output ({detail_text})"


def write_results_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    column: (
                        stringify_usage(row.get(column))
                        if column == "usage"
                        else row.get(column, "")
                    )
                    for column in CSV_COLUMNS
                }
            )


def average(values: list[float]) -> float | None:
    return sum(values) / len(values) if values else None


def result_number_values(rows: list[dict[str, Any]], key: str) -> list[int]:
    values: list[int] = []
    for row in rows:
        value = token_count_value(row.get(key))
        if isinstance(value, int):
            values.append(value)
    return values


def summarize_results(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_model: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_model.setdefault(row.get("model_key", ""), []).append(row)

    summaries: list[dict[str, Any]] = []
    for model_key, model_rows in sorted(by_model.items()):
        latencies = [
            float(row["latency_seconds"])
            for row in model_rows
            if row.get("latency_seconds") not in (None, "")
        ]
        total_tokens = result_number_values(model_rows, "total_tokens")
        summaries.append(
            {
                "model_key": model_key,
                "model_name": model_rows[0].get("model_name", ""),
                "provider_model": model_rows[0].get("provider_model", ""),
                "attempted": len(model_rows),
                "completed": sum(
                    1
                    for row in model_rows
                    if row.get("response") or row.get("output_files") or row.get("output_urls")
                ),
                "errors": sum(1 for row in model_rows if row.get("error")),
                "avg_latency_seconds": average(latencies),
                "total_tokens": sum(total_tokens) if total_tokens else None,
                "avg_total_tokens": average(total_tokens),
            }
        )
    return summaries


def write_sheet_rows(ws: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def style_header(ws: Any) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def set_widths(ws: Any, widths: dict[int, int]) -> None:
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width


def save_workbook_atomic(workbook: Any, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_name(f"{output_path.stem}.tmp{output_path.suffix}")
    workbook.save(temp_path)
    temp_path.replace(output_path)


def populate_results_sheets(
    wb: Any,
    rows: list[dict[str, Any]],
    skipped: list[SkippedTest],
) -> None:
    summary_ws = wb.create_sheet("Run Summary", 0)
    result_ws = wb.create_sheet("Run Results", 1)
    skipped_ws = wb.create_sheet("Skipped Tests", 2)

    write_sheet_rows(
        summary_ws,
        [
            [
                "Model Key",
                "Model Name",
                "Provider Model",
                "Attempted",
                "Completed",
                "Errors",
                "Average Latency Seconds",
                "Total Tokens",
                "Average Total Tokens",
            ]
        ],
    )
    for summary in summarize_results(rows):
        summary_ws.append(
            [
                summary["model_key"],
                summary["model_name"],
                summary["provider_model"],
                summary["attempted"],
                summary["completed"],
                summary["errors"],
                summary["avg_latency_seconds"],
                summary["total_tokens"],
                summary["avg_total_tokens"],
            ]
        )

    result_ws.append(CSV_COLUMNS)
    for row in rows:
        result_ws.append(
            [
                stringify_usage(row.get(column)) if column == "usage" else row.get(column, "")
                for column in CSV_COLUMNS
            ]
        )

    skipped_ws.append(["Test ID", "Category", "Criterion", "Eval Method", "Reason"])
    for test in skipped:
        skipped_ws.append(
            [test.test_id, test.category, test.criterion, test.eval_method, test.reason]
        )

    for ws in [summary_ws, result_ws, skipped_ws]:
        style_header(ws)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    set_widths(
        summary_ws,
        {1: 16, 2: 28, 3: 34, 4: 12, 5: 12, 6: 10, 7: 22, 8: 14, 9: 20},
    )
    set_widths(
        result_ws,
        {
            1: 20,
            2: 22,
            3: 16,
            4: 26,
            5: 16,
            6: 34,
            7: 12,
            8: 24,
            9: 28,
            10: 10,
            11: 20,
            14: 14,
            15: 80,
            16: 58,
            17: 58,
            18: 16,
            19: 14,
            20: 17,
            21: 17,
            22: 14,
            23: 42,
            24: 14,
        },
    )
    set_widths(skipped_ws, {1: 12, 2: 26, 3: 30, 4: 22, 5: 38})


def write_live_results_workbook(
    output_path: Path,
    rows: list[dict[str, Any]],
    skipped: list[SkippedTest],
) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    populate_results_sheets(wb, rows, skipped)
    save_workbook_atomic(wb, output_path)


def write_results_workbook(
    source_workbook: Path,
    output_path: Path,
    rows: list[dict[str, Any]],
    skipped: list[SkippedTest],
) -> None:
    wb = load_workbook(source_workbook)
    for sheet_name in ["Run Summary", "Run Results", "Skipped Tests"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    populate_results_sheets(wb, rows, skipped)
    save_workbook_atomic(wb, output_path)


def result_row(
    run_id: str,
    model: dict[str, Any],
    test: PromptTest,
    provider: str,
    rubric_fingerprints: dict[str, str] | None = None,
    output_type: str = "text",
    response: str = "",
    output_files: list[str] | None = None,
    output_urls: list[str] | None = None,
    latency_seconds: float | None = None,
    usage: dict[str, Any] | None = None,
    error: str = "",
) -> dict[str, Any]:
    usage_data = usage or {}
    rubric_fingerprint = (rubric_fingerprints or {}).get(test.test_id, "")
    return {
        "run_id": run_id,
        "timestamp": dt.datetime.now(dt.timezone.utc).isoformat(),
        "model_key": model.get("key", ""),
        "model_name": model.get("name", model.get("key", "")),
        "provider": provider,
        "provider_model": model.get("model", ""),
        "test_id": test.test_id,
        "category": test.category,
        "criterion": test.criterion,
        "weight": test.weight,
        "eval_method": test.eval_method,
        "prompt_source": test.prompt_source,
        "input_source": test.input_source,
        "output_type": output_type,
        "response": response,
        "output_files": "\n".join(output_files or []),
        "output_urls": "\n".join(output_urls or []),
        "latency_seconds": round(latency_seconds, 3) if latency_seconds is not None else "",
        "prompt_tokens": usage_token_count(usage_data, "prompt_tokens"),
        "completion_tokens": usage_token_count(usage_data, "completion_tokens"),
        "reasoning_tokens": reasoning_token_count(usage_data),
        "total_tokens": usage_token_count(usage_data, "total_tokens"),
        "usage": usage_data,
        "error": error,
        "prompt_fingerprint": prompt_fingerprint(test),
        "rubric_fingerprint": rubric_fingerprint,
        "benchmark_fingerprint": benchmark_fingerprint(model, test, rubric_fingerprints or {}),
        "cache_status": "fresh",
        "cache_source": "",
    }


def print_plan(
    config: dict[str, Any],
    tests: list[PromptTest],
    skipped: list[SkippedTest],
    models: list[dict[str, Any]],
    skipped_scored_models: list[dict[str, Any]] | None = None,
    planned_pair_count: int | None = None,
    reused_pair_count: int = 0,
    scored_model_filter_ignored: bool = False,
    pending_pairs: list[tuple[dict[str, Any], PromptTest]] | None = None,
    only_changed_tests: bool = False,
    changed_tests: set[str] | None = None,
) -> None:
    skipped_scored_models = skipped_scored_models or []
    print(f"Models enabled: {len(models)}")
    for model in models:
        capabilities = ", ".join(sorted(model_capabilities(config, model)))
        print(
            f"  - {model.get('key')}: {model.get('name')} "
            f"({model.get('model')}; {capabilities})"
        )
    if skipped_scored_models:
        print(f"Already scored models skipped: {len(skipped_scored_models)}")
        for model in skipped_scored_models:
            print(f"  - {model.get('key')}: {model.get('name')}")
    if scored_model_filter_ignored:
        print("Already scored model filter: ignored because prompt-level result reuse is enabled")
    print(f"Automated tests: {len(tests)}")
    by_category: dict[str, int] = {}
    for test in tests:
        by_category[test.category] = by_category.get(test.category, 0) + 1
    for category, count in sorted(by_category.items()):
        print(f"  - {category}: {count}")
    print(f"Skipped tests: {len(skipped)}")
    for test in skipped:
        print(f"  - {test.test_id}: {test.reason}")
    unsupported = unsupported_tests(config, models, tests)
    if unsupported:
        print("Selected tests without an enabled compatible model:")
        for test in unsupported:
            required = "image" if is_image_test(test) else "text"
            print(f"  - {test.test_id}: needs {required} capability")
    if reused_pair_count:
        print(f"Matching prior results reused: {reused_pair_count}")
    if only_changed_tests:
        changed_tests = changed_tests or set()
        changed_text = ", ".join(sorted(changed_tests)) if changed_tests else "(none)"
        print(f"Only changed prompts: on ({changed_text})")
    api_calls = planned_pair_count
    if api_calls is None:
        api_calls = len(planned_pairs(config, models, tests))
    print(f"Total API calls for selected model/test pairs: {api_calls}")
    if pending_pairs and api_calls:
        by_test: dict[str, int] = {}
        by_model: dict[str, int] = {}
        for model, test in pending_pairs:
            by_test[test.test_id] = by_test.get(test.test_id, 0) + 1
            model_key = clean_text(model.get("key"))
            by_model[model_key] = by_model.get(model_key, 0) + 1

        print("Pending API calls by Test ID:")
        for test_id, count in sorted(by_test.items()):
            print(f"  - {test_id}: {count}")

        print("Pending API calls by model:")
        for model_key, count in sorted(by_model.items()):
            print(f"  - {model_key}: {count}")


def print_parallel_plan(
    config: dict[str, Any],
    pairs: list[tuple[dict[str, Any], PromptTest]],
    parallel_products: bool,
    requested_workers: int,
) -> None:
    groups = group_pairs_by_product(config, pairs)
    workers = product_worker_count(parallel_products, requested_workers, len(groups))
    mode = "parallel" if parallel_products and workers > 1 else "serial"
    print(f"Product lanes: {len(groups)} ({mode}; workers: {workers})")
    for lane, lane_pairs in groups.items():
        model_keys = []
        seen_models: set[str] = set()
        for model, _test in lane_pairs:
            model_key = clean_text(model.get("key"))
            if model_key and model_key not in seen_models:
                seen_models.add(model_key)
                model_keys.append(model_key)
        print(
            f"  - {product_lane_label(lane)}: "
            f"{len(model_keys)} model(s), {len(lane_pairs)} call(s)"
        )


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, help="Path to the source .xlsx prompt workbook.")
    parser.add_argument(
        "--rubric-workbook",
        help="Optional grading rubric workbook used to verify selected Test IDs are scoreable.",
    )
    parser.add_argument(
        "--rubric-sheet",
        help="Rubric sheet name. Defaults to Scoring Guide/Rubric when present.",
    )
    parser.add_argument(
        "--allow-missing-rubric",
        action="store_true",
        help="Warn instead of failing when selected prompt Test IDs are absent from the rubric workbook.",
    )
    parser.add_argument("--config", help="Optional path to model_eval_models.json.")
    parser.add_argument(
        "--models-workbook",
        help="Path to an .xlsx model list. First sheet must contain at least a Model ID column.",
    )
    parser.add_argument("--sheet", default="Test Prompts", help="Prompt-library sheet name.")
    parser.add_argument("--output-dir", help="Output directory. Defaults to outputs/model_tests/<timestamp>.")
    parser.add_argument("--dry-run", action="store_true", help="Print the run plan without API calls.")
    parser.add_argument("--force", action="store_true", help="Ignore existing responses.jsonl in output-dir.")
    parser.add_argument(
        "--max-tokens",
        type=int,
        help="Override max output tokens per text response. Lower this if provider credit limits reject requests.",
    )
    parser.add_argument(
        "--reasoning-effort",
        choices=["none", "minimal", "low", "medium", "high", "xhigh"],
        help=(
            "OpenRouter reasoning effort override for models that support thinking tokens. "
            "Use none or minimal when token-cap failures leave blank/truncated visible output."
        ),
    )
    parser.add_argument(
        "--excel-every",
        type=int,
        default=1,
        help="Refresh responses.xlsx every N completed rows. Use 0 to disable live Excel output.",
    )
    parser.add_argument("--limit", type=int, help="Limit the number of selected prompt tests.")
    parser.add_argument("--only-tests", help="Comma-separated Test IDs, e.g. W1,C1,S2.")
    parser.add_argument("--only-models", help="Comma-separated model keys from the config.")
    parser.add_argument(
        "--website-seed-csv",
        help="Website db/seeds/model_variants.csv used to detect models that already have scores.",
    )
    parser.add_argument(
        "--skip-scored-models",
        action="store_true",
        help="Skip model rows whose model_id_string already has at least one website score.",
    )
    parser.add_argument(
        "--reuse-matching-results",
        action="store_true",
        help=(
            "Reuse prior successful responses for the same model, prompt/input, "
            "and rubric row instead of making another API call."
        ),
    )
    parser.add_argument(
        "--only-changed-tests",
        action="store_true",
        help=(
            "When reusing results, only call the API for Test IDs whose prompt/input "
            "content is new or changed compared with previous model-test workbooks. "
            "This avoids backfilling old missing/error pairs."
        ),
    )
    parser.add_argument(
        "--history-dir",
        help=(
            "Directory containing previous model-test output folders with responses.jsonl. "
            "Defaults to the parent of --output-dir when provided."
        ),
    )
    parser.add_argument("--include-image", action="store_true", help="Include image-generation prompts.")
    parser.add_argument("--include-evidence", action="store_true", help="Include evidence/privacy/security rows.")
    parser.add_argument("--include-manual-review", action="store_true", help="Include manual reviewer rows.")
    parser.add_argument(
        "--rate-limit-skip-after",
        type=int,
        default=3,
        help=(
            "Skip remaining tests for a model after this many rate-limit errors. "
            "Use 0 to keep trying every selected pair."
        ),
    )
    parser.add_argument(
        "--parallel-products",
        action="store_true",
        help=(
            "Run different product/provider lanes concurrently while keeping each "
            "lane's models and tests in series."
        ),
    )
    parser.add_argument(
        "--product-workers",
        type=int,
        default=0,
        help=(
            "Maximum concurrent product lanes when --parallel-products is set. "
            "Use 0 to allow one worker per product lane."
        ),
    )
    parser.add_argument(
        "--verbose-errors",
        action="store_true",
        help="Print full tracebacks for failed model calls.",
    )
    parser.add_argument(
        "--no-inherit-shorthand",
        action="store_true",
        help="Disable workbook shorthand inheritance for blank prompts and ^ input cells.",
    )
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    workbook_path = Path(args.workbook).expanduser()
    rubric_workbook_path = (
        Path(args.rubric_workbook).expanduser() if args.rubric_workbook else None
    )
    config_path = Path(args.config).expanduser() if args.config else None
    models_workbook_path = (
        Path(args.models_workbook).expanduser() if args.models_workbook else None
    )
    requested_output_dir = Path(args.output_dir).expanduser() if args.output_dir else None
    requested_website_seed_csv = (
        Path(args.website_seed_csv).expanduser() if args.website_seed_csv else None
    )
    requested_history_dir = Path(args.history_dir).expanduser() if args.history_dir else None

    if config_path:
        config = load_json(config_path)
    else:
        config = default_config()
    if models_workbook_path:
        config = read_model_workbook(models_workbook_path, config)
    elif not config.get("models"):
        raise ValueError("Provide --models-workbook or a --config file with models.")
    if args.max_tokens is not None:
        if args.max_tokens < 1:
            raise ValueError("--max-tokens must be greater than 0.")
        config.setdefault("request", {})["max_tokens"] = args.max_tokens
    if args.reasoning_effort:
        config.setdefault("request", {})["reasoning_effort"] = args.reasoning_effort
    if args.excel_every < 0:
        raise ValueError("--excel-every must be 0 or greater.")
    if args.rate_limit_skip_after < 0:
        raise ValueError("--rate-limit-skip-after must be 0 or greater.")
    if args.product_workers < 0:
        raise ValueError("--product-workers must be 0 or greater.")

    models = enabled_models(config, args.only_models)
    skipped_scored_models: list[dict[str, Any]] = []
    scored_model_filter_ignored = bool(args.skip_scored_models and args.reuse_matching_results)
    if args.skip_scored_models and not args.reuse_matching_results:
        if not requested_website_seed_csv:
            raise ValueError("--skip-scored-models requires --website-seed-csv.")
        if not requested_website_seed_csv.exists():
            raise ValueError(f"Website seed CSV not found: {requested_website_seed_csv}")
        scored_model_keys = read_scored_model_keys(requested_website_seed_csv)
        models, skipped_scored_models = filter_scored_models(models, scored_model_keys)
    tests = read_prompt_library(
        workbook_path=workbook_path,
        sheet_name=args.sheet,
        inherit_shorthand=not args.no_inherit_shorthand,
    )
    selected_tests, skipped = eligible_tests(tests, args)
    rubric_fingerprints: dict[str, str] = {}
    if rubric_workbook_path:
        if not rubric_workbook_path.exists():
            raise ValueError(f"Rubric workbook not found: {rubric_workbook_path}")
        rubric_fingerprints = read_rubric_test_fingerprints(
            rubric_workbook_path,
            args.rubric_sheet,
        )
        missing_rubric_test_ids = validate_rubric_coverage(
            selected_tests,
            set(rubric_fingerprints),
            args.allow_missing_rubric,
        )
        if missing_rubric_test_ids:
            preview = ", ".join(missing_rubric_test_ids[:25])
            suffix = (
                f" ... and {len(missing_rubric_test_ids) - 25} more"
                if len(missing_rubric_test_ids) > 25
                else ""
            )
            print(
                "Rubric coverage warning: "
                f"{len(missing_rubric_test_ids)} selected Test ID(s) missing "
                f"from rubric ({preview}{suffix})."
            )

    pairs = planned_pairs(config, models, selected_tests)
    history_dir = requested_history_dir
    if history_dir is None and requested_output_dir is not None:
        history_dir = requested_output_dir.parent
    if history_dir is None:
        history_dir = Path("outputs") / "model_tests"
    pair_fingerprints = {
        (clean_text(model.get("key")), test.test_id): benchmark_fingerprint(
            model,
            test,
            rubric_fingerprints,
        )
        for model, test in pairs
    }
    prompt_fingerprints_by_test_id = {
        test.test_id: prompt_fingerprint(test)
        for test in selected_tests
    }
    historical_prompt_fingerprints = (
        history_prompt_fingerprints(history_dir, requested_output_dir)
        if args.reuse_matching_results
        else {}
    )
    changed_prompt_test_ids = (
        changed_test_ids(
            selected_tests,
            prompt_fingerprints_by_test_id,
            historical_prompt_fingerprints,
        )
        if args.reuse_matching_results
        else set()
    )
    model_identity_by_key = {
        clean_text(model.get("key")): (
            clean_text(model.get("model")),
            clean_text(model.get("provider")),
        )
        for model in models
    }
    cached_rows_by_fingerprint = (
        read_cached_result_rows(
            history_dir=history_dir,
            output_dir=requested_output_dir,
            needed_fingerprints=set(pair_fingerprints.values()),
            pair_fingerprints=pair_fingerprints,
            prompt_fingerprints_by_test_id=prompt_fingerprints_by_test_id,
            rubric_fingerprints=rubric_fingerprints,
            model_identity_by_key=model_identity_by_key,
        )
        if args.reuse_matching_results
        else {}
    )
    cached_pairs = {
        pair_key: cached_rows_by_fingerprint[fingerprint]
        for pair_key, fingerprint in pair_fingerprints.items()
        if fingerprint in cached_rows_by_fingerprint
    }
    current_output_rows = (
        []
        if args.force or requested_output_dir is None
        else load_existing_jsonl(requested_output_dir / "responses.jsonl")
    )
    current_completed_pairs = matching_existing_pair_keys(
        current_output_rows,
        pair_fingerprints,
    )
    api_pairs = [
        (model, test)
        for model, test in pairs
        if (clean_text(model.get("key")), test.test_id) not in cached_pairs
        and (clean_text(model.get("key")), test.test_id) not in current_completed_pairs
        and (not args.only_changed_tests or test.test_id in changed_prompt_test_ids)
    ]

    if args.dry_run:
        print_plan(
            config,
            selected_tests,
            skipped,
            models,
            skipped_scored_models,
            planned_pair_count=len(api_pairs),
            reused_pair_count=len(cached_pairs),
            scored_model_filter_ignored=scored_model_filter_ignored,
            pending_pairs=api_pairs,
            only_changed_tests=args.only_changed_tests,
            changed_tests=changed_prompt_test_ids,
        )
        print_parallel_plan(config, api_pairs, args.parallel_products, args.product_workers)
        return 0
    if not models:
        raise ValueError("No enabled models found in the config.")
    if not selected_tests:
        raise ValueError("No eligible tests selected.")
    if not pairs:
        raise ValueError("No eligible model/test pairs. Check model capabilities in the config.")
    validate_api_keys(config, api_pairs)
    validate_openrouter_model_ids(config, api_pairs)

    print_plan(
        config,
        selected_tests,
        skipped,
        models,
        skipped_scored_models,
        planned_pair_count=len(api_pairs),
        reused_pair_count=len(cached_pairs),
        scored_model_filter_ignored=scored_model_filter_ignored,
        pending_pairs=api_pairs,
        only_changed_tests=args.only_changed_tests,
        changed_tests=changed_prompt_test_ids,
    )
    print_parallel_plan(config, api_pairs, args.parallel_products, args.product_workers)

    output_dir = make_output_dir(requested_output_dir)
    run_id = output_dir.name
    jsonl_path = output_dir / "responses.jsonl"
    live_xlsx_path = output_dir / "responses.xlsx"
    existing_rows = [] if args.force else load_existing_jsonl(jsonl_path)
    planned_keys = {(clean_text(model.get("key")), test.test_id) for model, test in pairs}
    api_planned_keys = {
        (clean_text(model.get("key")), test.test_id) for model, test in api_pairs
    }
    existing_pair_keys = matching_existing_pair_keys(existing_rows, pair_fingerprints)
    completed = existing_pair_keys & api_planned_keys
    all_rows = list(existing_rows)
    pair_order = {
        (clean_text(model.get("key")), test.test_id): index
        for index, (model, test) in enumerate(pairs)
    }

    reused_now = 0
    for model, test in pairs:
        pair_key = (clean_text(model.get("key")), test.test_id)
        cached_row = cached_pairs.get(pair_key)
        if not cached_row or pair_key in existing_pair_keys:
            continue
        row = reusable_result_row(
            cached_row,
            run_id=run_id,
            cache_source=clean_text(cached_row.get("_cache_source_path")),
        )
        append_jsonl(jsonl_path, row)
        all_rows.append(row)
        existing_pair_keys.add(pair_key)
        reused_now += 1
    if reused_now:
        print(f"Reused {reused_now} matching prior result(s).", flush=True)

    def ordered_rows() -> list[dict[str, Any]]:
        return sorted(
            all_rows,
            key=lambda row: (
                pair_order.get(
                    (clean_text(row.get("model_key")), clean_text(row.get("test_id"))),
                    len(pair_order),
                ),
                clean_text(row.get("timestamp")),
            ),
        )

    sleep_seconds = float(config.get("request", {}).get("sleep_seconds", 0.5))
    if args.excel_every:
        write_live_results_workbook(live_xlsx_path, ordered_rows(), skipped)

    total = len(api_pairs)
    done_count = len(completed)
    progress_lock = threading.Lock()
    output_lock = threading.Lock()
    token_warning_lock = threading.Lock()
    token_budget_warning_lines: list[str] = []

    def next_progress() -> int:
        nonlocal done_count
        with progress_lock:
            done_count += 1
            return done_count

    def persist_row(row: dict[str, Any]) -> None:
        with output_lock:
            append_jsonl(jsonl_path, row)
            all_rows.append(row)
            if args.excel_every and len(all_rows) % args.excel_every == 0:
                write_live_results_workbook(live_xlsx_path, ordered_rows(), skipped)

    def record_token_budget_warning(row: dict[str, Any]) -> None:
        warning = token_budget_warning(row)
        if not warning:
            return
        line = (
            "TOKEN WARNING: "
            f"{clean_text(row.get('model_key'))} / {clean_text(row.get('test_id'))}: "
            f"{warning}"
        )
        with token_warning_lock:
            token_budget_warning_lines.append(line)
        print(f"  {line}", flush=True)

    def run_pair(model: dict[str, Any], test: PromptTest) -> dict[str, Any]:
        provider = model.get("provider", "")
        started = time.monotonic()
        if is_image_test(test):
            response, output_files, output_urls, usage = call_image_model(
                config, model, test, output_dir
            )
            return result_row(
                run_id=run_id,
                model=model,
                test=test,
                provider=provider,
                rubric_fingerprints=rubric_fingerprints,
                output_type="image",
                response=response,
                output_files=output_files,
                output_urls=output_urls,
                latency_seconds=time.monotonic() - started,
                usage=usage,
            )

        response, usage = call_openai_compatible(config, model, create_messages(test))
        return result_row(
            run_id=run_id,
            model=model,
            test=test,
            provider=provider,
            rubric_fingerprints=rubric_fingerprints,
            response=response,
            latency_seconds=time.monotonic() - started,
            usage=usage,
        )

    def run_lane(
        lane: str,
        lane_pairs: list[tuple[dict[str, Any], PromptTest]],
    ) -> None:
        rate_limit_errors_by_model: dict[str, int] = {}
        rate_limited_models: set[str] = set()
        lane_label = product_lane_label(lane)
        show_lane = args.parallel_products and len(product_groups) > 1

        for model, test in lane_pairs:
            provider = model.get("provider", "")
            model_key = clean_text(model.get("key"))
            key = (model_key, test.test_id)
            if key in completed:
                print(f"Skipping existing {key[0]} / {key[1]}")
                continue
            if model_key in rate_limited_models:
                print(f"Skipping rate-limited {key[0]} / {key[1]}")
                continue

            current = next_progress()
            lane_prefix = f"[{lane_label}] " if show_lane else ""
            print(f"[{current}/{total}] {lane_prefix}{model_key} -> {test.test_id}", flush=True)
            started = time.monotonic()
            try:
                row = run_pair(model, test)
            except Exception as exc:  # pragma: no cover - depends on remote APIs
                if isinstance(exc, RateLimitError):
                    rate_limit_errors_by_model[model_key] = (
                        rate_limit_errors_by_model.get(model_key, 0) + 1
                    )
                    if (
                        args.rate_limit_skip_after
                        and rate_limit_errors_by_model[model_key] >= args.rate_limit_skip_after
                    ):
                        rate_limited_models.add(model_key)
                row = result_row(
                    run_id=run_id,
                    model=model,
                    test=test,
                    provider=provider,
                    rubric_fingerprints=rubric_fingerprints,
                    output_type="image" if is_image_test(test) else "text",
                    latency_seconds=time.monotonic() - started,
                    error=redact_secrets(f"{type(exc).__name__}: {exc}"),
                )
                if args.verbose_errors:
                    print(redact_secrets(traceback.format_exc()), file=sys.stderr)
                else:
                    print(f"  error: {row['error']}", file=sys.stderr)
                if model_key in rate_limited_models:
                    print(
                        "  rate limit: skipping remaining tests for "
                        f"{model_key} after {args.rate_limit_skip_after} rate-limit errors",
                        file=sys.stderr,
                    )

            persist_row(row)
            record_token_budget_warning(row)
            if sleep_seconds:
                time.sleep(sleep_seconds)

    product_groups = group_pairs_by_product(config, api_pairs)
    workers = product_worker_count(args.parallel_products, args.product_workers, len(product_groups))
    if workers > 1:
        print(f"Running product lanes in parallel with {workers} workers.", flush=True)
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
            future_to_lane = {
                executor.submit(run_lane, lane, lane_pairs): lane
                for lane, lane_pairs in product_groups.items()
            }
            for future in concurrent.futures.as_completed(future_to_lane):
                lane = future_to_lane[future]
                try:
                    future.result()
                except Exception as exc:
                    raise RuntimeError(f"Product lane {product_lane_label(lane)} failed.") from exc
    else:
        for lane, lane_pairs in product_groups.items():
            run_lane(lane, lane_pairs)

    final_rows = ordered_rows()
    if args.excel_every:
        write_live_results_workbook(live_xlsx_path, final_rows, skipped)
    write_results_csv(output_dir / "responses.csv", final_rows)
    write_results_workbook(
        source_workbook=workbook_path,
        output_path=output_dir / "model_test_results.xlsx",
        rows=final_rows,
        skipped=skipped,
    )

    if token_budget_warning_lines:
        print("Token budget warnings:", flush=True)
        for line in token_budget_warning_lines:
            print(f"  {line}", flush=True)

    print(f"Wrote {output_dir / 'responses.jsonl'}")
    print(f"Wrote {live_xlsx_path}")
    print(f"Wrote {output_dir / 'responses.csv'}")
    print(f"Wrote {output_dir / 'model_test_results.xlsx'}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv[1:]))
    except KeyboardInterrupt:
        print("Cancelled.", file=sys.stderr)
        raise SystemExit(130)
    except Exception as exc:
        print(f"Error: {redact_secrets(exc)}", file=sys.stderr)
        raise SystemExit(1)
